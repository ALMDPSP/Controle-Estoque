"""
Controle de Estoque - Entrada e Saída de Equipamentos
------------------------------------------------------
Programa em Python (Flask) com interface web (HTTP).

- Localmente: guarda os dados num banco SQLite (estoque.db), sem
  precisar configurar nada.
- No Render (produção): usa o PostgreSQL do próprio Render, através
  da variável de ambiente DATABASE_URL.

Tem login por usuário/senha e um botão para exportar os dados para
Excel (.xlsx) a qualquer momento.

Como rodar localmente:
    pip install -r requirements.txt
    python app.py

Depois abra no navegador:
    http://localhost:5000
    usuário: admin  senha: admin123  (troque depois de entrar)
"""

import io
import os
import unicodedata
import zipfile
import secrets
import hmac
import time
import base64
import hashlib
import struct
import json
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, request, jsonify, render_template, redirect,
    url_for, session, send_file, flash,
)
from werkzeug.security import check_password_hash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from cryptography.fernet import Fernet, InvalidToken
import qrcode

import db

app = Flask(__name__)
APP_BUILD = "2026-09-04-mfa-layout-simplificado-v38"
app.secret_key = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("SESSION_COOKIE_SECURE", "0") == "1",
)

# Proteções leves de autenticação. O limite é mantido em memória do processo
# para não exigir serviços externos e não altera nenhuma API já existente.
LOGIN_ATTEMPTS = {}
LOGIN_MAX_FAILURES = 5
LOGIN_WINDOW_SECONDS = 10 * 60

def _client_ip():
    forwarded = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    return forwarded or request.remote_addr or "desconhecido"

def _login_key(username):
    return f"{_client_ip()}|{(username or '').strip().lower()}"

def _prune_login_attempts(key):
    agora = time.time()
    tentativas = [t for t in LOGIN_ATTEMPTS.get(key, []) if agora - t < LOGIN_WINDOW_SECONDS]
    if tentativas:
        LOGIN_ATTEMPTS[key] = tentativas
    else:
        LOGIN_ATTEMPTS.pop(key, None)
    return tentativas

def _login_wait_seconds(username):
    key = _login_key(username)
    tentativas = _prune_login_attempts(key)
    if len(tentativas) < LOGIN_MAX_FAILURES:
        return 0
    return max(1, int(LOGIN_WINDOW_SECONDS - (time.time() - tentativas[0])))

def _register_login_failure(username):
    key = _login_key(username)
    tentativas = _prune_login_attempts(key)
    tentativas.append(time.time())
    LOGIN_ATTEMPTS[key] = tentativas

def _clear_login_failures(username):
    LOGIN_ATTEMPTS.pop(_login_key(username), None)

def _csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token

def _csrf_ok():
    recebido = request.form.get("csrf_token", "") or request.headers.get("X-CSRF-Token", "")
    esperado = session.get("_csrf_token", "")
    return bool(recebido and esperado and hmac.compare_digest(recebido, esperado))

@app.context_processor
def _inject_security_helpers():
    return {"csrf_token": _csrf_token()}


# ---------------------------------------------------------------------
# MFA / TOTP — Microsoft Authenticator e Google Authenticator
# ---------------------------------------------------------------------

MFA_ISSUER = os.environ.get("MFA_ISSUER", "Controle de Estoque")
MFA_ATTEMPTS = {}
MFA_MAX_FAILURES = 5
MFA_WINDOW_SECONDS = 5 * 60


def _mfa_cipher():
    material = os.environ.get("MFA_ENCRYPTION_KEY") or app.secret_key
    chave = base64.urlsafe_b64encode(hashlib.sha256(material.encode("utf-8")).digest())
    return Fernet(chave)


def _protect_mfa_secret(secret):
    return _mfa_cipher().encrypt(secret.encode("utf-8")).decode("ascii")


def _unprotect_mfa_secret(token):
    if not token:
        return None
    try:
        return _mfa_cipher().decrypt(token.encode("ascii")).decode("utf-8")
    except (InvalidToken, ValueError, TypeError):
        return None


def _base32_secret():
    return base64.b32encode(secrets.token_bytes(20)).decode("ascii").rstrip("=")


def _totp_code(secret, timestamp=None, interval=30, digits=6):
    timestamp = time.time() if timestamp is None else timestamp
    contador = int(timestamp // interval)
    padded = secret + "=" * ((8 - len(secret) % 8) % 8)
    key = base64.b32decode(padded.upper(), casefold=True)
    digest = hmac.new(key, struct.pack(">Q", contador), hashlib.sha1).digest()
    offset = digest[-1] & 0x0F
    valor = struct.unpack(">I", digest[offset:offset + 4])[0] & 0x7FFFFFFF
    return str(valor % (10 ** digits)).zfill(digits)


def _verify_totp(secret, codigo, window=1):
    codigo = "".join(ch for ch in str(codigo or "") if ch.isdigit())
    if len(codigo) != 6 or not secret:
        return False
    agora = time.time()
    return any(hmac.compare_digest(_totp_code(secret, agora + passo * 30), codigo) for passo in range(-window, window + 1))


def _mfa_uri(username, secret):
    from urllib.parse import quote
    label = quote(f"{MFA_ISSUER}:{username}")
    issuer = quote(MFA_ISSUER)
    return f"otpauth://totp/{label}?secret={secret}&issuer={issuer}&algorithm=SHA1&digits=6&period=30"


def _qr_data_uri(texto):
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=7, border=3)
    qr.add_data(texto)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def _gerar_codigos_recuperacao(qtd=8):
    alfabeto = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    codigos = []
    for _ in range(qtd):
        bruto = "".join(secrets.choice(alfabeto) for _ in range(10))
        codigos.append(bruto[:5] + "-" + bruto[5:])
    return codigos


def _hash_recovery_code(codigo):
    normalizado = "".join(ch for ch in str(codigo or "").upper() if ch.isalnum())
    return hashlib.sha256(normalizado.encode("utf-8")).hexdigest()


def _consume_recovery_code(usuario, codigo):
    if not usuario:
        return False
    try:
        hashes = json.loads(usuario.get("mfa_recovery_codes") or "[]")
    except Exception:
        hashes = []
    alvo = _hash_recovery_code(codigo)
    if alvo not in hashes:
        return False
    hashes.remove(alvo)
    db.atualizar_codigos_recuperacao_mfa(usuario["id"], json.dumps(hashes))
    return True


def _mfa_key(username):
    return f"{_client_ip()}|{(username or '').strip().lower()}"


def _mfa_prune(key):
    agora = time.time()
    tentativas = [t for t in MFA_ATTEMPTS.get(key, []) if agora - t < MFA_WINDOW_SECONDS]
    if tentativas:
        MFA_ATTEMPTS[key] = tentativas
    else:
        MFA_ATTEMPTS.pop(key, None)
    return tentativas


def _mfa_wait_seconds(username):
    tentativas = _mfa_prune(_mfa_key(username))
    if len(tentativas) < MFA_MAX_FAILURES:
        return 0
    return max(1, int(MFA_WINDOW_SECONDS - (time.time() - tentativas[0])))


def _mfa_register_failure(username):
    key = _mfa_key(username)
    tentativas = _mfa_prune(key)
    tentativas.append(time.time())
    MFA_ATTEMPTS[key] = tentativas


def _mfa_clear_failures(username):
    MFA_ATTEMPTS.pop(_mfa_key(username), None)


def _set_pending_login(usuario, proximo):
    csrf = session.get("_csrf_token") or secrets.token_urlsafe(32)
    session.clear()
    session["_csrf_token"] = csrf
    session["pending_user_id"] = usuario["id"]
    session["pending_username"] = usuario["username"]
    session["pending_role"] = usuario["role"]
    session["pending_precisa_trocar_senha"] = usuario.get("precisa_trocar_senha") == "1"
    session["pending_next"] = proximo or url_for("dashboard")


def _start_password_change_before_mfa(usuario, proximo):
    """Cria uma sessão restrita apenas à troca da senha no primeiro acesso.

    A senha já foi validada, mas o login ainda não é considerado concluído
    enquanto o usuário não criar a senha pessoal e finalizar o MFA.
    """
    csrf = session.get("_csrf_token") or secrets.token_urlsafe(32)
    session.clear()
    session["_csrf_token"] = csrf
    session["user_id"] = usuario["id"]
    session["username"] = usuario["username"]
    session["role"] = usuario["role"]
    session["precisa_trocar_senha"] = True
    session["primeiro_acesso_mfa_pendente"] = True
    session["primeiro_acesso_next"] = proximo or url_for("dashboard")


def _finalize_login(usuario=None):
    if usuario is None:
        user_id = session.get("pending_user_id")
        usuario = db.buscar_usuario_por_id(user_id) if user_id else None
    if not usuario:
        session.clear()
        return redirect(url_for("login"))
    precisa_trocar = usuario.get("precisa_trocar_senha") == "1"
    proximo = session.get("pending_next") or url_for("dashboard")
    session.clear()
    session["_csrf_token"] = secrets.token_urlsafe(32)
    session["user_id"] = usuario["id"]
    session["username"] = usuario["username"]
    session["role"] = usuario["role"]
    session["precisa_trocar_senha"] = precisa_trocar
    db.registrar_evento_login(usuario["username"], _client_ip(), "sucesso", "login concluído com MFA" if usuario.get("mfa_enabled") == "1" else "login realizado")
    if precisa_trocar:
        return redirect(url_for("trocar_senha"))
    return redirect(proximo)


# ---------------------------------------------------------------------
# Autenticação
# ---------------------------------------------------------------------

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", proximo=request.path))
        if session.get("precisa_trocar_senha") and request.endpoint not in ("trocar_senha", "logout"):
            return redirect(url_for("trocar_senha"))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", proximo=request.path))
        if session.get("precisa_trocar_senha"):
            return redirect(url_for("trocar_senha"))
        if session.get("role") != "admin":
            return jsonify({"erro": "Apenas administradores podem fazer isso."}), 403
        return view(*args, **kwargs)
    return wrapped


def admin_page_required(view):
    """Proteção para páginas administrativas com retorno amigável ao usuário."""
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", proximo=request.path))
        if session.get("precisa_trocar_senha"):
            return redirect(url_for("trocar_senha"))
        if session.get("role") != "admin":
            return redirect(url_for("dashboard", acesso_admin="1"))
        return view(*args, **kwargs)
    return wrapped


def role_required(*roles):
    permitidos=set(roles)
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login", proximo=request.path))
            if session.get("precisa_trocar_senha"):
                return redirect(url_for("trocar_senha"))
            role=session.get("role") or "user"
            # compatibilidade: perfis antigos 'user' funcionam como operador
            if role == "user": role = "operador"
            if role not in permitidos:
                return jsonify({"erro": "Seu perfil não possui permissão para esta operação."}), 403
            return view(*args, **kwargs)
        return wrapped
    return decorator


def edit_required(view):
    return role_required("admin", "gestor", "operador")(view)


def manager_required(view):
    return role_required("admin", "gestor")(view)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", erro=None, username_value="")

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")

    if not _csrf_ok():
        db.registrar_evento_login(username, _client_ip(), "falha", "token de segurança inválido")
        return render_template("login.html", erro="A sessão de segurança expirou. Atualize a página e tente novamente.", username_value=username), 400

    espera = _login_wait_seconds(username)
    if espera > 0:
        minutos = max(1, (espera + 59) // 60)
        db.registrar_evento_login(username, _client_ip(), "bloqueado", "limite de tentativas excedido")
        return render_template(
            "login.html",
            erro=f"Muitas tentativas de acesso. Aguarde aproximadamente {minutos} minuto(s) e tente novamente.",
            username_value=username,
        ), 429

    usuario = db.buscar_usuario_por_username(username)

    if not usuario or not check_password_hash(usuario["password_hash"], password):
        _register_login_failure(username)
        db.registrar_evento_login(username, _client_ip(), "falha", "usuário ou senha inválidos")
        return render_template("login.html", erro="Usuário ou senha inválidos.", username_value=username), 401

    _clear_login_failures(username)
    proximo = request.args.get("proximo") or url_for("dashboard")

    # Primeiro acesso:
    # senha temporária -> criação da senha pessoal -> configuração/validação do MFA -> sistema.
    if usuario.get("precisa_trocar_senha") == "1":
        _start_password_change_before_mfa(usuario, proximo)
        db.registrar_evento_login(
            usuario["username"],
            _client_ip(),
            "primeiro_acesso_senha_pendente",
            "senha temporária validada; aguardando criação da senha pessoal antes do MFA",
        )
        return redirect(url_for("trocar_senha"))

    # Demais acessos: MFA é obrigatório para todos os perfis.
    _set_pending_login(usuario, proximo)
    if usuario.get("mfa_enabled") == "1":
        db.registrar_evento_login(usuario["username"], _client_ip(), "mfa_pendente", "senha validada; aguardando segundo fator")
        return redirect(url_for("mfa_verificar"))
    db.registrar_evento_login(usuario["username"], _client_ip(), "mfa_configuracao_exigida", "usuário deve ativar MFA obrigatório")
    return redirect(url_for("mfa_configurar"))


@app.route("/mfa/verificar", methods=["GET", "POST"])
def mfa_verificar():
    user_id = session.get("pending_user_id")
    if not user_id:
        return redirect(url_for("login"))
    usuario = db.buscar_usuario_por_id(user_id)
    if not usuario:
        session.clear()
        return redirect(url_for("login"))
    if usuario.get("mfa_enabled") != "1":
        return redirect(url_for("mfa_configurar"))

    erro = None
    if request.method == "POST":
        if not _csrf_ok():
            erro = "A sessão de segurança expirou. Atualize a página e tente novamente."
            return render_template("mfa_verificar.html", username=usuario["username"], erro=erro), 400

        espera = _mfa_wait_seconds(usuario["username"])
        if espera > 0:
            minutos = max(1, (espera + 59) // 60)
            db.registrar_evento_login(usuario["username"], _client_ip(), "mfa_bloqueado", "limite de tentativas MFA excedido")
            erro = f"Muitas tentativas de MFA. Aguarde aproximadamente {minutos} minuto(s)."
            return render_template("mfa_verificar.html", username=usuario["username"], erro=erro), 429

        codigo = (request.form.get("codigo") or "").strip()
        secret = _unprotect_mfa_secret(usuario.get("mfa_secret"))
        totp_ok = _verify_totp(secret, codigo) if secret else False
        recovery_ok = False
        if not totp_ok and len("".join(ch for ch in codigo if ch.isalnum())) >= 8:
            recovery_ok = _consume_recovery_code(usuario, codigo)

        if not (totp_ok or recovery_ok):
            _mfa_register_failure(usuario["username"])
            db.registrar_evento_login(usuario["username"], _client_ip(), "mfa_falha", "código MFA inválido")
            erro = "Código inválido. Informe o código de 6 dígitos do Authenticator ou um código de recuperação."
            return render_template("mfa_verificar.html", username=usuario["username"], erro=erro), 401

        _mfa_clear_failures(usuario["username"])
        db.registrar_evento_login(usuario["username"], _client_ip(), "mfa_validado", "código de recuperação utilizado" if recovery_ok else "código TOTP validado")
        return _finalize_login(usuario)

    return render_template("mfa_verificar.html", username=usuario["username"], erro=erro)


@app.route("/mfa/configurar", methods=["GET", "POST"])
def mfa_configurar():
    pending = bool(session.get("pending_user_id"))
    user_id = session.get("pending_user_id") or session.get("user_id")
    if not user_id:
        return redirect(url_for("login"))
    usuario = db.buscar_usuario_por_id(user_id)
    if not usuario:
        session.clear()
        return redirect(url_for("login"))

    # Todos os perfis entram por aqui obrigatoriamente no primeiro acesso sem MFA.
    obrigatorio = pending
    if usuario.get("mfa_enabled") == "1":
        return redirect(url_for("mfa_verificar") if pending else url_for("pagina_seguranca"))

    setup_key = f"mfa_setup_secret_{user_id}"
    secret = session.get(setup_key)
    if not secret:
        secret = _base32_secret()
        session[setup_key] = secret
    qr_uri = _mfa_uri(usuario["username"], secret)
    qr_data = _qr_data_uri(qr_uri)
    erro = None

    if request.method == "POST":
        if not _csrf_ok():
            erro = "A sessão de segurança expirou. Atualize a página e tente novamente."
            return render_template("mfa_configurar.html", username=usuario["username"], secret=secret, qr_data=qr_data, erro=erro, obrigatorio=obrigatorio), 400
        codigo = request.form.get("codigo", "")
        if not _verify_totp(secret, codigo):
            db.registrar_evento_login(usuario["username"], _client_ip(), "mfa_config_falha", "código de confirmação inválido")
            erro = "O código não confere. Aguarde um novo código no Authenticator e tente novamente."
            return render_template("mfa_configurar.html", username=usuario["username"], secret=secret, qr_data=qr_data, erro=erro, obrigatorio=obrigatorio), 400

        codigos = _gerar_codigos_recuperacao()
        hashes = [_hash_recovery_code(c) for c in codigos]
        db.salvar_mfa_usuario(user_id, _protect_mfa_secret(secret), json.dumps(hashes))
        session.pop(setup_key, None)
        session["mfa_recovery_codes_once"] = codigos
        session["mfa_recovery_username"] = usuario["username"]
        session["mfa_setup_pending"] = pending
        db.registrar_evento_login(usuario["username"], _client_ip(), "mfa_ativado", "MFA TOTP ativado")
        return redirect(url_for("mfa_codigos_recuperacao"))

    return render_template("mfa_configurar.html", username=usuario["username"], secret=secret, qr_data=qr_data, erro=erro, obrigatorio=obrigatorio)


@app.route("/mfa/codigos-recuperacao")
def mfa_codigos_recuperacao():
    codigos = session.get("mfa_recovery_codes_once")
    if not codigos:
        if session.get("pending_user_id"):
            return redirect(url_for("mfa_verificar"))
        if session.get("user_id"):
            return redirect(url_for("pagina_seguranca"))
        return redirect(url_for("login"))
    return render_template(
        "mfa_recuperacao.html",
        username=session.get("mfa_recovery_username") or session.get("pending_username") or session.get("username"),
        codigos=codigos,
        pending=bool(session.get("mfa_setup_pending")),
    )


@app.route("/mfa/concluir", methods=["POST"])
def mfa_concluir():
    if not _csrf_ok():
        return redirect(url_for("mfa_codigos_recuperacao"))
    session.pop("mfa_recovery_codes_once", None)
    session.pop("mfa_recovery_username", None)
    pending = bool(session.pop("mfa_setup_pending", False))
    if pending and session.get("pending_user_id"):
        usuario = db.buscar_usuario_por_id(session.get("pending_user_id"))
        return _finalize_login(usuario)
    return redirect(url_for("pagina_seguranca") if session.get("user_id") else url_for("login"))


@app.route("/seguranca")
@login_required
def pagina_seguranca():
    usuario = db.buscar_usuario_por_id(session.get("user_id"))
    return render_template(
        "seguranca.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
        usuario=usuario,
    )


@app.route("/mfa/desativar", methods=["POST"])
@login_required
def mfa_desativar():
    usuario = db.buscar_usuario_por_id(session.get("user_id"))
    if not usuario or usuario.get("mfa_enabled") != "1":
        return redirect(url_for("pagina_seguranca"))
    flash("O MFA é obrigatório para todos os perfis e não pode ser desativado. Em caso de troca ou perda do aparelho, solicite a um Administrador o reset do MFA.", "erro")
    return redirect(url_for("pagina_seguranca"))


@app.route("/trocar-senha", methods=["GET", "POST"])
@login_required
def trocar_senha():
    if not session.get("precisa_trocar_senha"):
        return redirect(url_for("dashboard"))

    if request.method == "GET":
        return render_template("trocar_senha.html", username=session.get("username"), erro=None)

    if not _csrf_ok():
        return render_template(
            "trocar_senha.html",
            username=session.get("username"),
            erro="A sessão de segurança expirou. Atualize a página e tente novamente.",
        ), 400

    nova = request.form.get("nova_senha", "")
    confirmar = request.form.get("confirmar_senha", "")

    if len(nova) < 6:
        return render_template("trocar_senha.html", username=session.get("username"),
                                erro="A senha precisa ter pelo menos 6 caracteres.")
    if nova != confirmar:
        return render_template("trocar_senha.html", username=session.get("username"),
                                erro="As senhas não conferem.")

    user_id = session["user_id"]
    username = session.get("username")
    primeiro_acesso_mfa = bool(session.get("primeiro_acesso_mfa_pendente"))
    proximo = session.get("primeiro_acesso_next") or url_for("dashboard")

    db.trocar_senha(user_id, nova)
    db.registrar_evento_login(
        username,
        _client_ip(),
        "senha_alterada",
        "senha pessoal criada no primeiro acesso; MFA será exigido em seguida" if primeiro_acesso_mfa else "senha atualizada pelo usuário",
    )

    if primeiro_acesso_mfa:
        usuario = db.buscar_usuario_por_id(user_id)
        if not usuario:
            session.clear()
            return redirect(url_for("login"))

        # A sessão de troca de senha é encerrada e volta a ser uma sessão
        # pendente de MFA. Assim não existe caminho para o Dashboard antes
        # da conclusão do segundo fator.
        _set_pending_login(usuario, proximo)
        if usuario.get("mfa_enabled") == "1":
            db.registrar_evento_login(
                usuario["username"],
                _client_ip(),
                "mfa_pendente",
                "senha pessoal criada; aguardando validação do MFA",
            )
            return redirect(url_for("mfa_verificar"))

        db.registrar_evento_login(
            usuario["username"],
            _client_ip(),
            "mfa_configuracao_exigida",
            "senha pessoal criada; aguardando configuração do MFA obrigatório",
        )
        return redirect(url_for("mfa_configurar"))

    session["precisa_trocar_senha"] = False
    session["_csrf_token"] = secrets.token_urlsafe(32)
    return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------
# Páginas
# ---------------------------------------------------------------------

@app.route("/")
@login_required
def pagina_inicial():
    return redirect(url_for("dashboard"))


@app.route("/estoque")
@login_required
def index():
    return render_template(
        "index.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
    )


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template(
        "dashboard.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
    )


@app.route("/historico")
@login_required
def pagina_historico():
    return render_template(
        "historico.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
    )


@app.route("/relatorios")
@login_required
def pagina_relatorios():
    return render_template(
        "relatorios.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
    )


@app.route("/usuarios")
@admin_required
def pagina_usuarios():
    return render_template(
        "usuarios.html",
        username=session.get("username"),
        role=session.get("role") or "admin",
        is_admin=True,
        usuarios=db.listar_usuarios(),
    )


@app.route("/api/movimentacoes-recentes")
@login_required
def api_movimentacoes_recentes():
    try:
        limite=max(1,min(int(request.args.get("limite", 80)),500))
    except (TypeError,ValueError):
        limite=80
    movs=db.listar_movimentacoes_recentes(limite)
    itens={str(x.get("id")):x for x in db.listar_itens()}
    imobs={str(x.get("id")):x for x in db.listar_imobilizados()}
    for m in movs:
        tabela=m.get("tabela") or "itens"
        if tabela == "sistema":
            m["codigo"] = "META LOJAS"
            m["descricao"] = "Meta do lote de inauguração"
        else:
            ref=(imobs if tabela=="imobilizados" else itens).get(str(m.get("item_id")), {})
            m["codigo"]=ref.get("codigo","")
            m["descricao"]=ref.get("descricao","")
    return jsonify(movs)


@app.route("/api/auditoria-login")
@admin_required
def api_auditoria_login():
    try:
        limite=max(1,min(int(request.args.get("limite", 200)),500))
    except (TypeError,ValueError):
        limite=200
    eventos=db.listar_eventos_login_recentes(limite)
    # Não expõe o IP na interface para usuários da aplicação.
    for evento in eventos:
        evento.pop("ip", None)
    return jsonify(eventos)


@app.route("/api/busca-global")
@login_required
def api_busca_global():
    termo=unicodedata.normalize("NFD", (request.args.get("q") or "").strip().lower())
    termo="".join(c for c in termo if unicodedata.category(c)!="Mn")
    if len(termo)<1:
        return jsonify([])
    resultados=[]
    campos=("codigo","descricao","nro_serie","nro_patrimonio","nro_imobilizado","localizacao","local","armazenagem","vd_loja","filial_destino","pedido","chamado")
    def combina(obj):
        alvo=" ".join(str(obj.get(c) or "") for c in campos).lower()
        alvo=unicodedata.normalize("NFD",alvo)
        alvo="".join(c for c in alvo if unicodedata.category(c)!="Mn")
        return termo in alvo
    for nome,lista in (("Estoque",db.listar_itens()),("Imobilizados",db.listar_imobilizados())):
        for obj in lista:
            if combina(obj):
                resultados.append({
                    "origem":nome,"id":obj.get("id"),"codigo":obj.get("codigo","") or "",
                    "descricao":obj.get("descricao","") or "","quantidade":obj.get("qtde","") or "",
                    "tipo_estoque":obj.get("tipo_estoque","") or "","status":obj.get("status","") or "",
                    "localizacao":obj.get("localizacao","") or "","nro_serie":obj.get("nro_serie","") or "",
                    "nro_patrimonio":obj.get("nro_patrimonio","") or "",
                    "filial_destino":obj.get("filial_destino","") or ""
                })
            if len(resultados)>=80: break
        if len(resultados)>=80: break
    for filial in db.listar_filiais(incluir_inativas=True):
        alvo=f"{filial.get('codigo','')} {filial.get('nome','')} {filial.get('cidade','')} {filial.get('uf','')}".lower()
        alvo=unicodedata.normalize("NFD",alvo); alvo="".join(c for c in alvo if unicodedata.category(c)!="Mn")
        if termo in alvo:
            resultados.append({"origem":"Filiais","id":filial.get("id"),"codigo":filial.get("codigo","") or "","descricao":filial.get("nome","") or "","quantidade":"","tipo_estoque":"","status":_status_filial_normalizado(filial.get("ativo")),"localizacao":f"{filial.get('cidade','')} / {filial.get('uf','')}","nro_serie":"","nro_patrimonio":"","filial_destino":""})
        if len(resultados)>=100: break
    for prod in db.listar_produtos():
        alvo=f"{prod.get('codigo','')} {prod.get('descricao','')}".lower()
        alvo=unicodedata.normalize("NFD",alvo)
        alvo="".join(c for c in alvo if unicodedata.category(c)!="Mn")
        if termo in alvo:
            resultados.append({"origem":"Cadastro de Produtos","id":prod.get("id"),"codigo":prod.get("codigo","") or "","descricao":prod.get("descricao","") or "","quantidade":"","tipo_estoque":"","status":"","localizacao":"","nro_serie":"","nro_patrimonio":"","filial_destino":""})
        if len(resultados)>=100: break
    return jsonify(resultados)


def _normalizar_exec(valor):
    texto=unicodedata.normalize("NFD",str(valor or "").strip().lower())
    return "".join(c for c in texto if unicodedata.category(c)!="Mn")

def _calcular_visao_executiva(itens=None, kit=None, filiais=None, meta=None):
    def _qtd_num(valor):
        try:return int(float(valor or 0))
        except Exception:return 0
    if itens is None:
        itens=db.obter_dashboard_compacto(1).get("itens", [])
    if kit is None:
        kit=db.listar_kit_padrao_loja()
    if filiais is None:
        filiais=db.listar_filiais(incluir_inativas=True)
    if meta is None:
        meta=db.obter_meta_lojas_expansao()
    expansao=[x for x in itens if _normalizar_exec(x.get("tipo_estoque"))=="expansao" and _qtd_num(x.get("qtde"))>0]
    req=[]
    for k in kit:
        necessario=max(1,int(k.get("quantidade") or 1))
        codigo_k=str(k.get("codigo") or "").strip()
        desc_k=_normalizar_exec(k.get("descricao"))
        disponivel=0
        for item in expansao:
            codigo_i=str(item.get("codigo") or "").strip()
            desc_i=_normalizar_exec(item.get("descricao"))
            combina=(codigo_k and codigo_i==codigo_k) or (desc_k and desc_i and (desc_k in desc_i or desc_i in desc_k))
            if combina:
                disponivel += _qtd_num(item.get("qtde"))
        lojas=disponivel//necessario
        req.append({"codigo":codigo_k,"descricao":k.get("descricao") or "","necessario":necessario,"disponivel":disponivel,"lojas":lojas})
    capacidade=min([x["lojas"] for x in req],default=0)
    planejadas=[f for f in filiais if _status_filial_normalizado(f.get("ativo"))=="inaugurar"]
    planejadas.sort(key=lambda f: (str(f.get("previsao_abertura") or "9999-99-99"), str(f.get("codigo") or "")))
    qtd_planejada=len(planejadas)
    atendiveis=min(capacidade,qtd_planejada)
    risco=max(0,qtd_planejada-capacidade)
    pct=100.0 if qtd_planejada==0 else min(100.0,(capacidade/qtd_planejada)*100.0)
    hoje=datetime.now().date()
    horizontes={}
    for dias in (30,60,90):
        limite=hoje+timedelta(days=dias)
        dentro=[]
        for f in planejadas:
            txt=str(f.get("previsao_abertura") or "").strip()
            try: dt=datetime.strptime(txt[:10],"%Y-%m-%d").date()
            except Exception: continue
            if hoje <= dt <= limite: dentro.append(f)
        qtd=len(dentro)
        horizontes[str(dias)]={"lojas":qtd,"atendiveis":min(capacidade,qtd),"risco":max(0,qtd-capacidade)}
    sem_data=sum(1 for f in planejadas if not str(f.get("previsao_abertura") or "").strip())
    deficits=[]
    for x in req:
        alvo=x["necessario"]*max(1,qtd_planejada or int(meta or 1))
        falta=max(0,alvo-x["disponivel"])
        if falta:
            deficits.append({**x,"falta":falta,"alvo":alvo})
    deficits.sort(key=lambda x:-x["falta"])
    return {
        "meta_lojas":int(meta or 10),"capacidade_lojas":capacidade,"lojas_a_inaugurar":qtd_planejada,
        "lojas_atendiveis":atendiveis,"lojas_em_risco":risco,"percentual_atendimento":round(pct,1),
        "itens_criticos":len(deficits),"estoque_expansao":sum(_qtd_num(x.get("qtde")) for x in expansao),
        "horizontes":horizontes,"sem_data":sem_data,"deficits":deficits[:8],
        "planejadas":[{"id":f.get("id"),"codigo":f.get("codigo"),"nome":f.get("nome"),"uf":f.get("uf"),"previsao_abertura":f.get("previsao_abertura"),"situacao":"ATENDIDA" if i<capacidade else "RISCO"} for i,f in enumerate(planejadas)]
    }

def _obter_ultimo_backup_info():
    """Retorna o último backup persistido no banco para todos os usuários."""
    try:
        usuario = db.obter_configuracao("ultimo_backup_usuario")
        data_hora = db.obter_configuracao("ultimo_backup_datahora")
        arquivo = db.obter_configuracao("ultimo_backup_arquivo")
        if data_hora:
            return {"usuario": usuario or "-", "data_hora": data_hora, "arquivo": arquivo or ""}
    except Exception:
        pass
    return None

@app.route("/api/status-sistema")
@login_required
def api_status_sistema():
    try:
        saude=db.obter_saude_sistema()
    except Exception as e:
        saude={"database":"indisponível","database_ok":False,"erro":str(e),"contagens":{},"inconsistencias":{"total":0}}
    saude.update({
        "ultimo_backup":_obter_ultimo_backup_info(),
        "perfil":session.get("role") or "user",
        "usuario":session.get("username"),
        "build":APP_BUILD,
    })
    return jsonify(saude)

@app.route("/api/visao-executiva")
@login_required
def api_visao_executiva():
    return jsonify(_calcular_visao_executiva())

@app.route("/api/dashboard-resumo")
@login_required
def api_dashboard_resumo():
    """Carga compacta do Dashboard em uma única chamada HTTP."""
    base=db.obter_dashboard_compacto(20)
    visao=_calcular_visao_executiva(
        itens=base.get("itens") or [],
        kit=base.get("kit") or [],
        filiais=base.get("filiais") or [],
        meta=base.get("meta_lojas") or 10,
    )
    try:
        status=db.obter_saude_sistema()
    except Exception as e:
        status={"database":"indisponível","database_ok":False,"erro":str(e),"contagens":{},"inconsistencias":{"total":0}}
    status.update({
        "ultimo_backup":_obter_ultimo_backup_info(),
        "build":APP_BUILD,
    })
    resposta=jsonify({
        "itens":base.get("itens") or [],
        "estoque_total":base.get("estoque_total") or 0,
        "imobilizados_total":base.get("imobilizados_total") or 0,
        "produtos":base.get("produtos") or [],
        "produtos_total":base.get("produtos_total") or 0,
        "kit":base.get("kit") or [],
        "filiais_ativas":base.get("filiais_ativas") or 0,
        "meta_lojas":base.get("meta_lojas") or 10,
        "movimentacoes":base.get("movimentacoes") or [],
        "status":status,
        "visao":visao,
    })
    resposta.headers["Cache-Control"]="private, max-age=5"
    return resposta

@app.route("/api/importacoes-recentes")
@admin_required
def api_importacoes_recentes():
    return jsonify(db.listar_importacoes_recentes(30))

@app.route("/gestao-dados")
@admin_page_required
def pagina_gestao_dados():
    return render_template("gestao_dados.html",username=session.get("username"),role=session.get("role") or "user",is_admin=session.get("role")=="admin")


UF_NOMES = {
    "AC":"Acre","AL":"Alagoas","AP":"Amapá","AM":"Amazonas","BA":"Bahia",
    "CE":"Ceará","DF":"Distrito Federal","ES":"Espírito Santo","GO":"Goiás",
    "MA":"Maranhão","MT":"Mato Grosso","MS":"Mato Grosso do Sul","MG":"Minas Gerais",
    "PA":"Pará","PB":"Paraíba","PR":"Paraná","PE":"Pernambuco","PI":"Piauí",
    "RJ":"Rio de Janeiro","RN":"Rio Grande do Norte","RS":"Rio Grande do Sul",
    "RO":"Rondônia","RR":"Roraima","SC":"Santa Catarina","SP":"São Paulo",
    "SE":"Sergipe","TO":"Tocantins",
}

MAPA_UF_POS = {
    "RR": (160, 75), "AP": (305, 85), "AM": (120, 165), "PA": (270, 170),
    "AC": (45, 230), "RO": (92, 220), "TO": (252, 225), "MA": (280, 175),
    "PI": (303, 194), "CE": (330, 190), "RN": (355, 195), "PB": (348, 205),
    "PE": (334, 219), "AL": (345, 233), "SE": (341, 248), "BA": (300, 255),
    "MT": (190, 233), "GO": (235, 262), "DF": (256, 256), "MS": (180, 296),
    "MG": (275, 282), "ES": (318, 288), "RJ": (307, 318), "SP": (257, 310),
    "PR": (241, 337), "SC": (248, 356), "RS": (224, 372),
}

STATUS_FILIAL_ROTULOS = {
    "ativa": "Ativa",
    "inaugurar": "Inaugurar",
    "pendente": "Pendente",
    "inativa": "Inativa",
}


def _cor_estado_pdf(estado):
    total = int(estado.get("total") or 0)
    dsp = int(estado.get("dsp") or 0)
    dpa = int(estado.get("dpa") or 0)
    if total <= 0:
        return colors.HexColor("#252D38"), None
    conhecidos = dsp + dpa
    if conhecidos <= 0:
        return colors.HexColor("#657083"), None
    if dsp > 0 and dpa == 0:
        return colors.HexColor("#3EA6FF"), None
    if dpa > 0 and dsp == 0:
        return colors.HexColor("#EF5260"), None
    return colors.HexColor("#3EA6FF"), colors.HexColor("#EF5260")


def _draw_round_label(pdf, x, y, w, h, title, value, fill="#1A2029", value_color="#FFFFFF"):
    pdf.setFillColor(colors.HexColor(fill))
    pdf.setStrokeColor(colors.HexColor("#2B3444"))
    pdf.roundRect(x, y, w, h, 10, stroke=1, fill=1)
    pdf.setFillColor(colors.HexColor("#8B96A8"))
    pdf.setFont("Helvetica", 8)
    pdf.drawString(x + 10, y + h - 14, title.upper())
    pdf.setFillColor(colors.HexColor(value_color))
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(x + 10, y + 12, str(value))


def _draw_store_icon(pdf, x, y, scale=1.0, accent="#3EA6FF"):
    aw = 34 * scale
    ah = 22 * scale
    pdf.setStrokeColor(colors.HexColor("#203040"))
    pdf.setLineWidth(1)
    pdf.setFillColor(colors.HexColor("#EAF4FF"))
    pdf.roundRect(x, y, aw, ah, 4 * scale, stroke=1, fill=1)
    pdf.setFillColor(colors.HexColor(accent))
    pdf.rect(x - 1 * scale, y + ah - 8 * scale, aw + 2 * scale, 8 * scale, stroke=0, fill=1)
    pdf.setFillColor(colors.white)
    stripe_w = (aw + 2 * scale) / 5.0
    for i in range(5):
        if i % 2 == 0:
            pdf.rect(x - 1 * scale + i * stripe_w, y + ah - 8 * scale, stripe_w, 8 * scale, stroke=0, fill=1)
    pdf.setFillColor(colors.HexColor("#D9E9F7"))
    pdf.rect(x + 4 * scale, y + 4 * scale, 8 * scale, 9 * scale, stroke=0, fill=1)
    pdf.setFillColor(colors.HexColor("#C8DDF0"))
    pdf.rect(x + 17 * scale, y + 4 * scale, 12 * scale, 12 * scale, stroke=0, fill=1)


def _draw_state_tile_pdf(pdf, cx, cy, estado, scale=1.0):
    w = 30 * scale
    h = 22 * scale
    x = cx - (w / 2.0)
    y = cy - (h / 2.0)
    c1, c2 = _cor_estado_pdf(estado)
    pdf.saveState()
    path = pdf.beginPath()
    path.roundRect(x, y, w, h, 4)
    pdf.clipPath(path, stroke=0, fill=0)
    if c2 is None:
        pdf.setFillColor(c1)
        pdf.rect(x, y, w, h, stroke=0, fill=1)
    else:
        conhecidos = max(1, int(estado.get("dsp") or 0) + int(estado.get("dpa") or 0))
        split = w * (int(estado.get("dsp") or 0) / conhecidos)
        pdf.setFillColor(c1)
        pdf.rect(x, y, split, h, stroke=0, fill=1)
        pdf.setFillColor(c2)
        pdf.rect(x + split, y, w - split, h, stroke=0, fill=1)
    pdf.restoreState()
    pdf.setStrokeColor(colors.HexColor("#E7F3FF"))
    pdf.setLineWidth(0.7)
    pdf.roundRect(x, y, w, h, 4, stroke=1, fill=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", max(6, 7.4 * scale))
    pdf.drawCentredString(cx, y + h - (7.8 * scale), str(estado.get("uf") or ""))
    pdf.setFont("Helvetica-Bold", max(4.5, 5.4 * scale))
    pdf.setFillColor(colors.HexColor("#E8FFF3"))
    pdf.drawRightString(cx - 1.5 * scale, y + 4.3 * scale, str(int(estado.get('ativa') or 0)))
    pdf.setFillColor(colors.HexColor("#FFFFFF"))
    pdf.drawCentredString(cx, y + 4.2 * scale, "|")
    pdf.setFillColor(colors.HexColor("#E0F1FF"))
    pdf.drawString(cx + 1.5 * scale, y + 4.3 * scale, str(int(estado.get('inaugurar') or 0)))


def _gerar_pdf_projecao_lojas(dados):
    estados = dados.get("estados") or []
    totais = dados.get("totais") or {}
    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=landscape(A4))
    larg, alt = landscape(A4)
    margem = 14 * mm
    area_w = larg - 2 * margem

    estados_validos = [e for e in estados if int(e.get("total") or 0) > 0]
    estados_ordenados = sorted(estados_validos, key=lambda e: (-int(e.get("total") or 0), e.get("estado") or ""))
    total_base = max(1, int(totais.get("total_geral") or 0) or sum(int(e.get("total") or 0) for e in estados_validos) or 1)
    top_estado = estados_ordenados[0] if estados_ordenados else None
    top5_total = sum(int(e.get("total") or 0) for e in estados_ordenados[:5])
    cobertura = len(estados_validos)
    taxa_ativas = (int(totais.get("ativa") or 0) / total_base) * 100.0

    def _footer(page_no):
        pdf.setStrokeColor(colors.HexColor("#253241"))
        pdf.line(margem, 10 * mm, larg - margem, 10 * mm)
        pdf.setFillColor(colors.HexColor("#8EA1B4"))
        pdf.setFont("Helvetica", 7.5)
        pdf.drawString(margem, 6.5 * mm, "© 2026 · Developed by Alexandre Martins · Relatório executivo de projeção de abertura e lojas")
        pdf.drawRightString(larg - margem, 6.5 * mm, f"Página {page_no}")

    def _panel(x, y, w, h, title=None, subtitle=None, radius=12):
        pdf.setFillColor(colors.HexColor("#151D27"))
        pdf.setStrokeColor(colors.HexColor("#2A3645"))
        pdf.roundRect(x, y, w, h, radius, stroke=1, fill=1)
        if title:
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 12)
            pdf.drawString(x + 12, y + h - 20, title)
        if subtitle:
            pdf.setFillColor(colors.HexColor("#91A7BD"))
            pdf.setFont("Helvetica", 8.2)
            pdf.drawString(x + 12, y + h - 33, subtitle)

    def _kpi_card(x, y, w, h, titulo, valor, cor, detalhe):
        pdf.setFillColor(colors.HexColor("#182230"))
        pdf.setStrokeColor(colors.HexColor("#314255"))
        pdf.roundRect(x, y, w, h, 11, stroke=1, fill=1)
        pdf.setFillColor(colors.HexColor("#9AB0C5"))
        pdf.setFont("Helvetica", 7.4)
        pdf.drawString(x + 10, y + h - 13, titulo.upper())
        pdf.setFillColor(colors.HexColor(cor))
        pdf.setFont("Helvetica-Bold", 17)
        pdf.drawString(x + 10, y + 19, str(valor))
        pdf.setFillColor(colors.HexColor("#7F95AA"))
        pdf.setFont("Helvetica", 6.8)
        pdf.drawString(x + 10, y + 8, detalhe)

    def _summary_row(x, y, w, label, value, color="#FFFFFF"):
        pdf.setFillColor(colors.HexColor("#1B2531"))
        pdf.roundRect(x, y, w, 22, 7, stroke=0, fill=1)
        pdf.setFillColor(colors.HexColor("#A0B5C9"))
        pdf.setFont("Helvetica", 8.2)
        pdf.drawString(x + 10, y + 7.5, label)
        pdf.setFillColor(colors.HexColor(color))
        pdf.setFont("Helvetica-Bold", 10.5)
        pdf.drawRightString(x + w - 10, y + 7.5, str(value))

    def _bullet_line(x, y, label, value, color="#DCE8F5"):
        pdf.setFillColor(colors.HexColor(color))
        pdf.circle(x + 3, y + 2.5, 2, stroke=0, fill=1)
        pdf.setFillColor(colors.HexColor("#DDE7F1"))
        pdf.setFont("Helvetica", 8)
        pdf.drawString(x + 10, y, label)
        pdf.setFont("Helvetica-Bold", 8.2)
        pdf.drawRightString(x + 198, y, str(value))

    def _state_bar_row(x, y, w, nome, total, ativa, inaugurar, pct, idx):
        h = 26
        fill = "#16212E" if idx % 2 == 0 else "#141D29"
        pdf.setFillColor(colors.HexColor(fill))
        pdf.roundRect(x, y, w, h, 7, stroke=0, fill=1)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 8.6)
        nome_curto = nome if len(nome) <= 28 else nome[:25] + "..."
        pdf.drawString(x + 8, y + 16, nome_curto)
        pdf.setFillColor(colors.HexColor("#8CA2B7"))
        pdf.setFont("Helvetica", 7)
        pdf.drawString(x + 8, y + 7, f"Total {total} · Ativas {ativa} · Inaugurar {inaugurar}")
        track_x = x + 150
        track_w = w - 215
        pdf.setFillColor(colors.HexColor("#0D1721"))
        pdf.roundRect(track_x, y + 8, track_w, 10, 5, stroke=0, fill=1)
        pdf.setFillColor(colors.HexColor("#4FB2FF"))
        pdf.roundRect(track_x, y + 8, max(6, track_w * max(0, min(1, pct / 100.0))), 10, 5, stroke=0, fill=1)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawRightString(x + w - 8, y + 11, f"{pct:.1f}%")

    def _mini_brand_card(x, y, w, h, titulo, valor, pct, color_hex):
        pdf.setFillColor(colors.HexColor("#1B2531"))
        pdf.roundRect(x, y, w, h, 8, stroke=0, fill=1)
        titulo_exib = titulo if len(str(titulo)) <= 10 else "Sem band."
        title_font = 7.0 if len(titulo_exib) > 8 else 7.8
        pdf.setFillColor(colors.HexColor("#8EA2B6"))
        pdf.setFont("Helvetica-Bold", title_font)
        pdf.drawCentredString(x + (w/2), y + h - 10, titulo_exib)
        pdf.setFillColor(colors.HexColor(color_hex))
        pdf.setFont("Helvetica-Bold", 10.8)
        pdf.drawString(x + 8, y + 14, str(valor))
        pdf.setFillColor(colors.HexColor("#D9E4EF"))
        pdf.setFont("Helvetica-Bold", 7.0)
        pdf.drawRightString(x + w - 8, y + 14, f"{pct:.1f}%")
        pdf.setFillColor(colors.HexColor("#0D1721"))
        pdf.roundRect(x + 8, y + 5, w - 16, 4, 2, stroke=0, fill=1)
        barra = (w - 16) * max(0, min(1, pct / 100.0))
        if barra > 0:
            pdf.setFillColor(colors.HexColor(color_hex))
            pdf.roundRect(x + 8, y + 5, max(6, barra), 4, 2, stroke=0, fill=1)

    def _pill(x, y, text, accent="#FFB648"):
        tw = pdf.stringWidth(text, "Helvetica-Bold", 7.2)
        w = tw + 18
        pdf.setFillColor(colors.HexColor("#101923"))
        pdf.setStrokeColor(colors.HexColor("#314255"))
        pdf.roundRect(x, y, w, 16, 8, stroke=1, fill=1)
        pdf.setFillColor(colors.HexColor(accent))
        pdf.setFont("Helvetica-Bold", 7.2)
        pdf.drawString(x + 9, y + 5, text)
        return w

    def _draw_page_bg():
        pdf.setFillColor(colors.HexColor("#0F1620"))
        pdf.rect(0, 0, larg, alt, stroke=0, fill=1)

    # ============================
    # Página 1 — Visão executiva
    # ============================
    _draw_page_bg()
    pdf.setTitle("Projecao de abertura e lojas")
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 21)
    pdf.drawString(margem, alt - margem, "Projeção de abertura e lojas")
    pdf.setFillColor(colors.HexColor("#9DB3C8"))
    pdf.setFont("Helvetica", 9)
    pdf.drawString(margem, alt - margem - 14, "Relatório executivo com visão consolidada das lojas cadastradas na aba Filiais.")
    pdf.drawRightString(larg - margem, alt - margem - 14, datetime.now().strftime("Gerado em %d/%m/%Y às %H:%M"))

    kpi_y = alt - margem - 66
    gap = 8
    card_w = (area_w - gap * 5) / 6.0
    card_h = 50
    kpis = [
        ("Total geral", totais.get("total_geral", 0), "#FFFFFF", "Base operacional"),
        ("Lojas ativas", totais.get("ativa", 0), "#52D69A", "Em operação"),
        ("A inaugurar", totais.get("inaugurar", 0), "#5AB4FF", "Planejadas"),
        ("Pendentes", totais.get("pendente", 0), "#FFBE55", "Aguardando definição"),
        ("DSP", totais.get("dsp", 0), "#63BAFF", "Bandeira azul"),
        ("DPA", totais.get("dpa", 0), "#FF7E88", "Bandeira vermelha"),
    ]
    for i, (titulo, valor, cor, detalhe) in enumerate(kpis):
        _kpi_card(margem + i * (card_w + gap), kpi_y, card_w, card_h, titulo, valor, cor, detalhe)

    content_y = 52
    content_h = kpi_y - 18 - content_y
    left_w = 510
    right_gap = 12
    right_x = margem + left_w + right_gap
    right_w = larg - margem - right_x

    # Painel esquerdo principal
    _panel(margem, content_y, left_w, content_h, "Participação percentual por estado", "Leitura dos estados com maior concentração de lojas na projeção.")
    pdf.setFillColor(colors.HexColor("#0F1924"))
    pdf.roundRect(margem + left_w - 126, content_y + content_h - 28, 112, 17, 7, stroke=0, fill=1)
    pdf.setFillColor(colors.HexColor("#E2ECF7"))
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawCentredString(margem + left_w - 70, content_y + content_h - 22, f"Base total: {total_base} loja(s)")

    pdf.setFillColor(colors.HexColor("#A5B8CA"))
    pdf.setFont("Helvetica", 8)
    pdf.drawString(margem + 12, content_y + content_h - 48, "Top 10 estados por participação no total de lojas.")
    row_y = content_y + content_h - 82
    row_h = 30
    for idx, e in enumerate(estados_ordenados[:10]):
        pct = (int(e.get("total") or 0) / total_base) * 100.0
        _state_bar_row(margem + 12, row_y - idx * row_h, left_w - 24, f"{e.get('estado')} ({e.get('uf')})", int(e.get("total") or 0), int(e.get("ativa") or 0), int(e.get("inaugurar") or 0), pct, idx)

    # Insights executivos no rodapé do painel esquerdo
    insight_y = content_y + 18
    insight_w = (left_w - 24 - 3 * 8) / 4.0
    top_inauguracao = max(estados_ordenados, key=lambda e: int(e.get("inaugurar") or 0), default=None)
    insights = [
        ("Cobertura nacional", f"{cobertura}/27", f"{(cobertura/27)*100:.1f}% das UFs com lojas", "#63BAFF"),
        ("Operação ativa", f"{taxa_ativas:.1f}%", f"{totais.get('ativa',0)} lojas ativas", "#52D69A"),
        ("Maior presença", f"{top_estado.get('uf') if top_estado else '-'} · {top_estado.get('total') if top_estado else 0}", f"{top_estado.get('estado') if top_estado else 'Sem dados'} lidera a base", "#FFBE55"),
        ("Maior inauguração", f"{top_inauguracao.get('uf') if top_inauguracao else '-'} · {top_inauguracao.get('inaugurar') if top_inauguracao else 0}", f"{top_inauguracao.get('estado') if top_inauguracao else 'Sem dados'} possui mais inaugurações", "#B197FC"),
    ]
    for i, (titulo, valor, detalhe, cor) in enumerate(insights):
        x = margem + 12 + i * (insight_w + 8)
        pdf.setFillColor(colors.HexColor("#182230"))
        pdf.roundRect(x, insight_y, insight_w, 58, 9, stroke=0, fill=1)
        pdf.setFillColor(colors.HexColor("#8FA3B8"))
        pdf.setFont("Helvetica", 7.2)
        pdf.drawString(x + 9, insight_y + 43, titulo.upper())
        pdf.setFillColor(colors.HexColor(cor))
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(x + 9, insight_y + 25, str(valor))
        pdf.setFillColor(colors.HexColor("#D7E2EE"))
        pdf.setFont("Helvetica", 7.2)
        pdf.drawString(x + 9, insight_y + 10, detalhe[:36])

    # Painel direito
    _panel(right_x, content_y, right_w, content_h)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(right_x + 12, content_y + content_h - 20, "Resumo executivo")

    cursor = content_y + content_h - 50
    for rot, val, cor in [
        ("Estados com lojas", cobertura, "#FFFFFF"),
        ("Ativas + inaugurar", int(totais.get("ativa", 0)) + int(totais.get("inaugurar", 0)), "#52D69A"),
        ("Pendentes", totais.get("pendente", 0), "#FFBE55"),
        ("Sem bandeira", totais.get("sem_bandeira", 0), "#A7B5C4"),
    ]:
        _summary_row(right_x + 12, cursor, right_w - 24, rot, val, cor)
        cursor -= 27

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(right_x + 12, cursor - 4, "Mensagem executiva")
    cursor -= 18
    _bullet_line(right_x + 12, cursor, "Base considerada", totais.get("total_geral", 0))
    cursor -= 14
    _bullet_line(right_x + 12, cursor, f"Estado líder: {top_estado.get('uf') if top_estado else '-'}", f"{(int(top_estado.get('total') or 0)/total_base*100):.1f}%" if top_estado else "0%", "#63BAFF")
    cursor -= 14
    _bullet_line(right_x + 12, cursor, "Cobertura nacional", f"{cobertura} UF(s)", "#52D69A")
    cursor -= 14
    _bullet_line(right_x + 12, cursor, "Taxa de operação ativa", f"{taxa_ativas:.1f}%", "#FFBE55")
    cursor -= 24

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(right_x + 12, cursor, "Top 5 estados")
    cursor -= 16
    for i, e in enumerate(estados_ordenados[:5], start=1):
        pct = (int(e.get("total") or 0) / total_base) * 100.0
        pdf.setFillColor(colors.HexColor("#1B2531"))
        pdf.roundRect(right_x + 12, cursor - 11, right_w - 24, 18, 6, stroke=0, fill=1)
        pdf.setFillColor(colors.HexColor("#DCE7F2"))
        pdf.setFont("Helvetica", 8)
        nome = f"{i}. {e.get('estado')} ({e.get('uf')})"
        pdf.drawString(right_x + 18, cursor, nome[:31])
        pdf.drawRightString(right_x + right_w - 18, cursor, f"{pct:.1f}% · {int(e.get('total') or 0)}")
        cursor -= 20

    cursor -= 2
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(right_x + 12, cursor, "Composição por bandeira")
    cursor -= 42
    total_marcas = max(1, int(totais.get("dsp", 0)) + int(totais.get("dpa", 0)) + int(totais.get("sem_bandeira", 0)))
    brand_w = (right_w - 24 - 2 * 8) / 3.0
    brands = [
        ("DSP", totais.get("dsp", 0), (int(totais.get("dsp", 0)) / total_marcas) * 100.0, "#63BAFF"),
        ("DPA", totais.get("dpa", 0), (int(totais.get("dpa", 0)) / total_marcas) * 100.0, "#FF7E88"),
        ("Sem bandeira", totais.get("sem_bandeira", 0), (int(totais.get("sem_bandeira", 0)) / total_marcas) * 100.0, "#9AAABA"),
    ]
    for i, (titulo, valor, pct, cor) in enumerate(brands):
        _mini_brand_card(right_x + 12 + i * (brand_w + 8), cursor, brand_w, 34, titulo, valor, pct, cor)

    _footer(1)

    # ============================
    # Página 2 — Detalhamento executivo
    # ============================
    pdf.showPage()
    _draw_page_bg()
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(margem, alt - margem, "Projeção detalhada por estado")
    pdf.setFillColor(colors.HexColor("#9DB3C8"))
    pdf.setFont("Helvetica", 9)
    pdf.drawString(margem, alt - margem - 14, "Tabela consolidada com percentual de participação, status operacional e distribuição por bandeira.")
    pdf.drawRightString(larg - margem, alt - margem - 14, datetime.now().strftime("Atualizado em %d/%m/%Y às %H:%M"))

    # Faixa de leitura executiva
    band_y = alt - margem - 64
    band_h = 48
    _panel(margem, band_y, area_w, band_h, radius=10)
    exec_msg = f"A projeção atual contempla {totais.get('total_geral',0)} loja(s), com {totais.get('ativa',0)} ativa(s), {totais.get('inaugurar',0)} a inaugurar e {totais.get('pendente',0)} pendente(s)."
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 11.5)
    pdf.drawString(margem + 12, band_y + 28, "Leitura executiva")
    pdf.setFillColor(colors.HexColor("#C7D5E3"))
    pdf.setFont("Helvetica", 8.6)
    pdf.drawString(margem + 12, band_y + 14, exec_msg)
    if top_estado:
        pdf.drawString(margem + 12, band_y + 4, f"Maior presença: {top_estado.get('estado')} ({top_estado.get('uf')}) com {top_estado.get('total')} loja(s), representando {(int(top_estado.get('total') or 0)/total_base)*100:.1f}% da base.")

    cols = [("UF", 24), ("Estado", 112), ("%", 36), ("Total", 40), ("Ativas", 44), ("Inaug.", 48), ("Pend.", 46), ("DSP", 36), ("DPA", 36), ("Sem", 40)]
    table_x = margem
    table_y = band_y - 26
    row_h = 18
    table_w = sum(w for _, w in cols)

    def _draw_table_header(ypos):
        pdf.setFillColor(colors.HexColor("#234C74"))
        pdf.roundRect(table_x, ypos, table_w, row_h, 5, stroke=0, fill=1)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 8)
        cx = table_x
        for title, w in cols:
            pdf.drawString(cx + 4, ypos + 6, title)
            cx += w

    _draw_table_header(table_y)
    y = table_y - row_h - 2
    alterna = False
    page_no = 2
    for e in estados_ordenados:
        if y < 18 * mm:
            _footer(page_no)
            pdf.showPage()
            page_no += 1
            _draw_page_bg()
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 15)
            pdf.drawString(margem, alt - margem, "Projeção detalhada por estado · continuação")
            pdf.setFillColor(colors.HexColor("#9DB3C8"))
            pdf.setFont("Helvetica", 8.5)
            pdf.drawString(margem, alt - margem - 14, "Continuação da tabela consolidada por estado.")
            table_y = alt - margem - 40
            _draw_table_header(table_y)
            y = table_y - row_h - 2
            alterna = False
        fill = "#19222D" if alterna else "#141D28"
        alterna = not alterna
        pdf.setFillColor(colors.HexColor(fill))
        pdf.roundRect(table_x, y, table_w, row_h, 3, stroke=0, fill=1)
        pct = (int(e.get("total") or 0) / total_base) * 100.0
        vals = [
            e.get("uf"), e.get("estado"), f"{pct:.1f}%", e.get("total"), e.get("ativa"), e.get("inaugurar"),
            e.get("pendente"), e.get("dsp"), e.get("dpa"), e.get("sem_bandeira")
        ]
        cx = table_x
        for idx, ((_, w), val) in enumerate(zip(cols, vals)):
            if idx == 2:
                pdf.setFillColor(colors.HexColor("#74BFFF"))
            elif idx == 4:
                pdf.setFillColor(colors.HexColor("#52D69A"))
            elif idx == 5:
                pdf.setFillColor(colors.HexColor("#63BAFF"))
            elif idx == 6:
                pdf.setFillColor(colors.HexColor("#FFBE55"))
            elif idx == 8:
                pdf.setFillColor(colors.HexColor("#FF7E88"))
            else:
                pdf.setFillColor(colors.HexColor("#DCE7F2"))
            pdf.setFont("Helvetica", 7.7)
            if idx >= 2:
                pdf.drawRightString(cx + w - 4, y + 6, str(val))
            else:
                label = str(val)
                if idx == 1 and len(label) > 23:
                    label = label[:20] + "..."
                pdf.drawString(cx + 4, y + 6, label)
            cx += w
        y -= row_h + 2

    _footer(page_no)
    pdf.save()
    buf.seek(0)
    return buf


def _status_filial_normalizado(valor):
    valor = str(valor or "").strip().lower()
    if valor in ("1", "ativa", "ativo", "true"):
        return "ativa"
    if valor == "inaugurar":
        return "inaugurar"
    if valor == "pendente":
        return "pendente"
    return "inativa"

def _dados_projecao_lojas():
    filiais = db.listar_filiais(incluir_inativas=True)
    por_uf = {}
    totais = {
        "ativa": 0, "inaugurar": 0, "pendente": 0, "inativa": 0,
        "dsp": 0, "dpa": 0, "sem_bandeira": 0, "total_geral": 0,
    }
    for f in filiais:
        status = _status_filial_normalizado(f.get("ativo"))
        bandeira = str(f.get("bandeira") or "").strip().upper()
        uf = str(f.get("uf") or "").strip().upper()[:2]
        totais[status] += 1
        if status != "inativa":
            totais["total_geral"] += 1
            if bandeira == "DSP":
                totais["dsp"] += 1
            elif bandeira == "DPA":
                totais["dpa"] += 1
            else:
                totais["sem_bandeira"] += 1
        if uf not in UF_NOMES:
            continue
        linha = por_uf.setdefault(uf, {
            "uf": uf, "estado": UF_NOMES[uf],
            "ativa": 0, "inaugurar": 0, "pendente": 0, "inativa": 0,
            "dsp": 0, "dpa": 0, "sem_bandeira": 0, "total": 0,
        })
        linha[status] += 1
        if status != "inativa":
            linha["total"] += 1
            if bandeira == "DSP":
                linha["dsp"] += 1
            elif bandeira == "DPA":
                linha["dpa"] += 1
            else:
                linha["sem_bandeira"] += 1
    estados = []
    for uf, nome in UF_NOMES.items():
        estados.append(por_uf.get(uf, {
            "uf": uf, "estado": nome,
            "ativa": 0, "inaugurar": 0, "pendente": 0, "inativa": 0,
            "dsp": 0, "dpa": 0, "sem_bandeira": 0, "total": 0,
        }))
    estados.sort(key=lambda x: (-x["total"], x["estado"]))
    return {"totais": totais, "estados": estados, "filiais": filiais}

@app.route("/api/projecao-lojas")
@login_required
def api_projecao_lojas():
    dados = _dados_projecao_lojas()
    return jsonify({"totais": dados["totais"], "estados": dados["estados"]})


def _preencher_planilha_dict(ws, dados, titulo=None):
    if titulo:
        ws.append([titulo])
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(1,len(dados[0]) if dados else 1))
        ws["A1"].font=Font(bold=True,size=14)
    if not dados:
        ws.append(["Sem dados"])
        return
    colunas=[]
    for obj in dados:
        for chave in obj.keys():
            if chave not in colunas: colunas.append(chave)
    ws.append(colunas)
    cab_row=ws.max_row
    fill=PatternFill("solid", fgColor="1F4E78")
    for c in ws[cab_row]:
        c.font=Font(bold=True,color="FFFFFF")
        c.fill=fill
        c.alignment=Alignment(horizontal="center")
    for obj in dados:
        ws.append([obj.get(c,"") for c in colunas])
    ws.freeze_panes=f"A{cab_row+1}"
    ws.auto_filter.ref=f"A{cab_row}:{get_column_letter(len(colunas))}{ws.max_row}"
    for idx,col in enumerate(colunas,1):
        largura=max(len(str(col)),12)
        for row in ws.iter_rows(min_row=cab_row+1,min_col=idx,max_col=idx):
            largura=max(largura,min(len(str(row[0].value or "")),45))
        ws.column_dimensions[get_column_letter(idx)].width=min(largura+2,48)


def _workbook_consolidado():
    wb=Workbook()
    wb.remove(wb.active)
    projecao = _dados_projecao_lojas()
    fontes=[
        ("Estoque",db.listar_itens()),
        ("Imobilizados",db.listar_imobilizados()),
        ("Produtos",db.listar_produtos()),
        ("Filiais",projecao["filiais"]),
        ("Projeção por UF",projecao["estados"]),
        ("Kit padrão",db.listar_kit_padrao_loja()),
        ("Movimentações",db.listar_todas_movimentacoes()),
    ]
    for nome,dados in fontes:
        ws=wb.create_sheet(nome[:31])
        _preencher_planilha_dict(ws,dados)
    ws_resumo=wb.create_sheet("Resumo Lojas")
    totais=projecao["totais"]
    _preencher_planilha_dict(ws_resumo,[{
        "Lojas ativas":totais["ativa"],
        "A inaugurar":totais["inaugurar"],
        "Pendentes":totais["pendente"],
        "Inativas":totais["inativa"],
        "Total geral":totais["total_geral"],
        "DSP":totais["dsp"],
        "DPA":totais["dpa"],
        "Sem bandeira":totais["sem_bandeira"],
    }])
    return wb


@app.route("/export-consolidado")
@login_required
def exportar_consolidado():
    wb=_workbook_consolidado()
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"relatorio_consolidado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/export-projecao-lojas")
@login_required
def exportar_projecao_lojas():
    dados = _dados_projecao_lojas()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo Geral"
    t = dados["totais"]
    resumo = [
        ("Indicador", "Quantidade"),
        ("Lojas ativas", t["ativa"]),
        ("Lojas a inaugurar", t["inaugurar"]),
        ("Lojas pendentes", t["pendente"]),
        ("Lojas inativas", t["inativa"]),
        ("Total geral", t["total_geral"]),
        ("DSP", t["dsp"]),
        ("DPA", t["dpa"]),
        ("Sem bandeira definida", t["sem_bandeira"]),
    ]
    for row in resumo:
        ws.append(row)
    for c in ws[1]:
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F4E78")
    ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=18

    ws_uf = wb.create_sheet("Por UF")
    _preencher_planilha_dict(ws_uf, dados["estados"], "Projeção de abertura e lojas por estado")

    ws_filiais = wb.create_sheet("Filiais")
    linhas=[]
    rotulos={"1":"Ativa","0":"Inativa","inaugurar":"Inaugurar","pendente":"Pendente"}
    for f in dados["filiais"]:
        x=dict(f)
        x["status"] = rotulos.get(str(f.get("ativo") or ""), str(f.get("ativo") or ""))
        linhas.append(x)
    _preencher_planilha_dict(ws_filiais, linhas, "Cadastro de filiais usado na projeção")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"projecao_abertura_lojas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export-projecao-lojas-pdf")
@login_required
def exportar_projecao_lojas_pdf():
    dados = _dados_projecao_lojas()
    pdf_buffer = _gerar_pdf_projecao_lojas(dados)
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"projecao_abertura_lojas_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
        mimetype="application/pdf",
    )


@app.route("/export-movimentacoes")
@login_required
def exportar_movimentacoes():
    inicio=(request.args.get("inicio") or "").strip()
    fim=(request.args.get("fim") or "").strip()
    movs=db.listar_todas_movimentacoes()
    if inicio:
        movs=[m for m in movs if str(m.get("data_hora") or "")[:10] >= inicio]
    if fim:
        movs=[m for m in movs if str(m.get("data_hora") or "")[:10] <= fim]
    itens={str(x.get("id")):x for x in db.listar_itens()}
    imobs={str(x.get("id")):x for x in db.listar_imobilizados()}
    for m in movs:
        tabela=m.get("tabela") or "itens"
        if tabela == "sistema":
            m["codigo"]="META LOJAS"
            m["descricao"]="Meta do lote de inauguração"
        else:
            ref=(imobs if tabela=="imobilizados" else itens).get(str(m.get("item_id")),{})
            m["codigo"]=ref.get("codigo","")
            m["descricao"]=ref.get("descricao","")
    wb=Workbook(); ws=wb.active; ws.title="Movimentações"; _preencher_planilha_dict(ws,movs)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    faixa=f"_{inicio or 'inicio'}_{fim or 'hoje'}"
    return send_file(buf,as_attachment=True,download_name=f"movimentacoes{faixa}.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/backup")
@login_required
def gerar_backup():
    agora = datetime.now()
    usuario = session.get("username") or "Usuário"
    data_hora = agora.strftime("%d/%m/%Y %H:%M")
    nome_arquivo = f"backup_controle_estoque_{agora.strftime('%Y%m%d_%H%M')}.zip"
    mem=io.BytesIO()
    with zipfile.ZipFile(mem,"w",zipfile.ZIP_DEFLATED) as z:
        wb=_workbook_consolidado()
        x=io.BytesIO(); wb.save(x); x.seek(0)
        z.writestr("estoque_backup.xlsx",x.read())
        if not db.IS_PG and os.path.exists(db.SQLITE_PATH):
            z.write(db.SQLITE_PATH,arcname="estoque.db")
        z.writestr("LEIA-ME.txt",f"Backup gerado em {data_hora} por {usuario}.\nContém Estoque, Imobilizados, Produtos, Filiais, Projeção por UF, Kit padrão e Histórico de movimentações.\n")

    # Registro persistente: permanece disponível após logout, novo login ou reinício da aplicação.
    db.salvar_configuracao("ultimo_backup_usuario", usuario, usuario)
    db.salvar_configuracao("ultimo_backup_datahora", data_hora, usuario)
    db.salvar_configuracao("ultimo_backup_arquivo", nome_arquivo, usuario)
    session["ultimo_backup"] = data_hora  # compatibilidade com versões anteriores

    mem.seek(0)
    resposta = send_file(mem,as_attachment=True,download_name=nome_arquivo,mimetype="application/zip")
    resposta.headers["X-Backup-Usuario"] = usuario
    resposta.headers["X-Backup-DataHora"] = data_hora
    return resposta


@app.route("/produtos")
@login_required
def pagina_produtos():
    return render_template("produtos.html", username=session.get("username"), role=session.get("role") or "user", is_admin=session.get("role") == "admin")


@app.route("/filiais")
@login_required
def pagina_filiais():
    return render_template("filiais.html", username=session.get("username"), role=session.get("role") or "user", is_admin=session.get("role") == "admin")


@app.route("/filiais/<int:filial_id>")
@login_required
def pagina_filial_detalhe(filial_id):
    filial=db.buscar_filial_por_id(filial_id)
    if not filial:
        return redirect(url_for("pagina_filiais"))
    codigo=str(filial.get("codigo") or "")
    itens=[x for x in db.listar_itens() if str(x.get("filial_destino") or "")==codigo]
    imobs=[x for x in db.listar_imobilizados() if str(x.get("filial_destino") or "")==codigo]
    visao=_calcular_visao_executiva()
    proj=next((x for x in visao.get("planejadas",[]) if int(x.get("id") or 0)==filial_id),None)
    return render_template("filial_detalhe.html",filial=filial,itens=itens,imobs=imobs,projecao=proj,username=session.get("username"),role=session.get("role") or "user",is_admin=session.get("role")=="admin")


@app.route("/projecao-lojas")
@login_required
def pagina_projecao_lojas():
    return render_template(
        "projecao_lojas.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
    )


@app.route("/leitor-codigo")
@login_required
def pagina_leitor_codigo():
    return render_template("leitor_codigo.html", username=session.get("username"), role=session.get("role") or "user", is_admin=session.get("role") == "admin")


@app.route("/acesso-celular")
@login_required
def pagina_acesso_celular():
    host = (request.host or "").split(":")[0].lower()
    porta = request.environ.get("SERVER_PORT") or "5000"
    if host in ("localhost", "127.0.0.1", "0.0.0.0"):
        ip = descobrir_ip_local()
        url_celular = f"http://{ip}:{porta}"
        modo = "rede_local"
    else:
        esquema = request.headers.get("X-Forwarded-Proto", request.scheme or "http").split(",")[0].strip()
        url_celular = f"{esquema}://{request.host}"
        modo = "publico"
    return render_template(
        "acesso_celular.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
        url_celular=url_celular,
        modo_acesso=modo,
    )


@app.route("/service-worker.js")
def service_worker():
    return send_file(
        os.path.join(app.static_folder, "service-worker.js"),
        mimetype="application/javascript",
        max_age=0,
    )


@app.route("/loja-virtual")
@login_required
def pagina_loja_virtual():
    return render_template("loja_virtual.html", username=session.get("username"), role=session.get("role") or "user", is_admin=session.get("role") == "admin")


@app.route("/api/filiais", methods=["GET"])
@login_required
def api_listar_filiais():
    incluir_inativas = request.args.get("inativas", "1") != "0"
    return jsonify(db.listar_filiais(incluir_inativas=incluir_inativas))


def _texto_celula_excel_filial(celula):
    """Converte a célula para texto, preservando zeros à esquerda quando possível."""
    valor = celula.value
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "1" if valor else "0"
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        fmt = str(celula.number_format or "")
        # Formatos como 0000 / 000000 preservam o código visual da loja.
        if isinstance(valor, int) or (isinstance(valor, float) and valor.is_integer()):
            inteiro = int(valor)
            apenas_zeros = fmt.replace(";", "").replace("@", "").strip()
            if apenas_zeros and set(apenas_zeros) <= {"0"}:
                return str(inteiro).zfill(len(apenas_zeros))
            return str(inteiro)
    return str(valor).strip()


def _cabecalho_filial_normalizado(valor):
    s = unicodedata.normalize("NFD", str(valor or ""))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    # Remove pontuação e separadores para aceitar cabeçalhos como
    # "Nome / identificação", exatamente como no Excel exportado pelo sistema.
    s = "".join(ch if ch.isalnum() else " " for ch in s.lower())
    s = " ".join(s.split())
    aliases = {
        "codigo": {
            "codigo", "codigo filial", "codigo da filial", "codigo loja", "codigo da loja",
            "numero loja", "numero da loja", "n loja", "loja", "filial",
        },
        "nome": {"nome", "nome identificacao", "nome filial", "nome da filial", "identificacao", "descricao", "descricao filial"},
        "cidade": {"cidade", "municipio"},
        "uf": {"uf", "estado", "sigla uf"},
        "bandeira": {"bandeira", "marca", "rede"},
        "status": {"status", "situacao", "situacao da loja", "ativo"},
        "previsao_abertura": {"previsao abertura", "previsao de abertura", "data abertura", "data de abertura", "abertura prevista"},
    }
    for campo, nomes in aliases.items():
        if s in nomes:
            return campo
    return None


def _status_filial_importacao(valor, atual=None):
    s = unicodedata.normalize("NFD", str(valor or "")).encode("ascii", "ignore").decode("ascii").strip().lower()
    if not s:
        return str(atual if atual is not None else "1")
    mapa = {
        "1": "1", "ativa": "1", "ativo": "1", "sim": "1", "true": "1",
        "0": "0", "inativa": "0", "inativo": "0", "nao": "0", "false": "0",
        "inaugurar": "inaugurar", "a inaugurar": "inaugurar", "inauguracao": "inaugurar",
        "pendente": "pendente", "pendencia": "pendente",
    }
    return mapa.get(s)

def _data_filial_iso(valor):
    if valor is None or valor == "":
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto=str(valor).strip()
    for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%Y/%m/%d"):
        try:
            return datetime.strptime(texto[:10],fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return texto[:10] if len(texto)>=10 else texto


@app.route("/export-filiais")
@login_required
def exportar_filiais_excel():
    filiais = db.listar_filiais(incluir_inativas=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Filiais"
    headers = ["Código da filial", "Nome / identificação", "Cidade", "UF", "Bandeira", "Status", "Previsão de abertura"]
    ws.append(headers)
    status_rotulos = {"1": "Ativa", "0": "Inativa", "inaugurar": "Inaugurar", "pendente": "Pendente"}
    for f in filiais:
        ws.append([
            str(f.get("codigo") or ""), f.get("nome") or "", f.get("cidade") or "",
            str(f.get("uf") or "").upper(), str(f.get("bandeira") or "").upper(),
            status_rotulos.get(str(f.get("ativo") or ""), str(f.get("ativo") or "")),
            str(f.get("previsao_abertura") or ""),
        ])

    cor_header = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = cor_header
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(1, ws.max_row)}"
    larguras = [20, 34, 24, 10, 14, 16, 20]
    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura
    for row in ws.iter_rows(min_row=2, max_col=7):
        row[0].number_format = "@"
        row[3].alignment = Alignment(horizontal="center")
        row[4].alignment = Alignment(horizontal="center")
        row[5].alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"filiais_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/filiais/importar/validar", methods=["POST"])
@manager_required
def api_validar_filiais_excel():
    arquivo=request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro":"Nenhum arquivo enviado."}),400
    if not arquivo.filename.lower().endswith((".xlsx",".xlsm")):
        return jsonify({"erro":"Envie um arquivo Excel (.xlsx ou .xlsm)."}),400
    try:
        wb=load_workbook(arquivo,read_only=True,data_only=True); ws=wb.active
        linhas_iter=ws.iter_rows(values_only=True)
        mapa_colunas={}; cab_row=0
        for numero_linha,valores in enumerate(linhas_iter,start=1):
            if numero_linha>10: break
            candidato={}
            for idx,valor in enumerate(valores or ()):
                campo=_cabecalho_filial_normalizado(valor)
                if campo and campo not in candidato: candidato[campo]=idx
            if "codigo" in candidato:
                mapa_colunas=candidato; cab_row=numero_linha; break
        if not mapa_colunas:
            return jsonify({"erro":"Não encontrei a coluna de código da filial."}),400
        atuais={str(f.get("codigo") or "").strip():f for f in db.listar_filiais(incluir_inativas=True)}
        vistos=set(); total=validas=ignoradas=criadas=atualizadas=sem_alteracao=0; erros=[]; amostra=[]
        for numero_linha,valores in enumerate(linhas_iter,start=cab_row+1):
            valores=tuple(valores or ())
            if not valores or all(v is None or str(v).strip()=="" for v in valores): continue
            total+=1
            def ler(campo):
                idx=mapa_colunas.get(campo)
                if idx is None or idx>=len(valores): return ""
                v=valores[idx]
                if v is None:return ""
                if isinstance(v,datetime):return v.strftime("%Y-%m-%d")
                if isinstance(v,float) and v.is_integer():return str(int(v))
                return str(v).strip()
            codigo=ler("codigo").strip(); uf=ler("uf").upper().strip(); bandeira=ler("bandeira").upper().strip(); status_txt=ler("status").strip()
            if not codigo or codigo in vistos:
                ignoradas+=1
                if len(erros)<15: erros.append(f"Linha {numero_linha}: código ausente ou repetido ({codigo or '-'}).")
                continue
            vistos.add(codigo)
            if uf and (len(uf)!=2 or uf not in UF_NOMES):
                ignoradas+=1
                if len(erros)<15:erros.append(f"Linha {numero_linha}: UF '{uf}' inválida.")
                continue
            if bandeira and bandeira not in ("DSP","DPA"):
                ignoradas+=1
                if len(erros)<15:erros.append(f"Linha {numero_linha}: bandeira '{bandeira}' inválida.")
                continue
            status="" if not status_txt else _status_filial_importacao(status_txt,None)
            if status_txt and status is None:
                ignoradas+=1
                if len(erros)<15:erros.append(f"Linha {numero_linha}: status '{status_txt}' inválido.")
                continue
            validas+=1
            atual=atuais.get(codigo)
            if atual:
                nome=ler("nome") or str(atual.get("nome") or ""); cidade=ler("cidade") or str(atual.get("cidade") or ""); nova_uf=uf or str(atual.get("uf") or "").upper(); nova_b=bandeira or str(atual.get("bandeira") or "").upper(); novo_s=status or str(atual.get("ativo") or "1"); nova_p=_data_filial_iso(ler("previsao_abertura")) or str(atual.get("previsao_abertura") or "")
                mudou=any([str(atual.get("nome") or "")!=nome,str(atual.get("cidade") or "")!=cidade,str(atual.get("uf") or "").upper()!=nova_uf,str(atual.get("bandeira") or "").upper()!=nova_b,str(atual.get("ativo") or "")!=novo_s,str(atual.get("previsao_abertura") or "")!=nova_p])
                if mudou: atualizadas+=1
                else: sem_alteracao+=1
            else: criadas+=1
            if len(amostra)<8: amostra.append({"codigo":codigo,"nome":ler("nome"),"uf":uf,"status":status or (atual or {}).get("ativo","1"),"acao":"Atualizar" if atual else "Criar"})
        return jsonify({"ok":True,"arquivo":arquivo.filename,"total_linhas":total,"validas":validas,"criadas":criadas,"atualizadas":atualizadas,"sem_alteracao":sem_alteracao,"ignoradas":ignoradas,"erros":erros,"amostra":amostra})
    except Exception as e:
        return jsonify({"erro":f"Erro ao validar a planilha: {e}"}),500

@app.route("/api/filiais/importar", methods=["POST"])
@manager_required
def api_importar_filiais_excel():
    # Mesmo padrão de segurança da importação da aba Estoque.
    senha = request.form.get("senha", "")
    usuario_atual = db.buscar_usuario_por_id(session["user_id"])
    if not usuario_atual or not check_password_hash(usuario_atual["password_hash"], senha):
        return jsonify({"erro": "Senha incorreta."}), 403

    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400
    if not arquivo.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"erro": "Envie um arquivo Excel (.xlsx ou .xlsm)."}), 400

    try:
        try:
            wb = load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return jsonify({"erro": "Não consegui abrir esse arquivo. Confirme se é um .xlsx válido."}), 400

        if ws is None:
            return jsonify({"erro": "A planilha não possui uma aba com dados."}), 400

        # Leitura sequencial, igual à importação do Estoque. Evita milhares de
        # acessos aleatórios ws.cell() no modo read_only, que é lento no Render.
        linhas_iter = ws.iter_rows(values_only=True)
        cabecalho = None
        mapa_colunas = {}
        cab_row = 0
        for numero_linha, valores in enumerate(linhas_iter, start=1):
            if numero_linha > 10:
                break
            candidato = {}
            for idx, valor in enumerate(valores or ()):
                campo = _cabecalho_filial_normalizado(valor)
                if campo and campo not in candidato:
                    candidato[campo] = idx
            if "codigo" in candidato:
                cabecalho = valores
                mapa_colunas = candidato
                cab_row = numero_linha
                break

        if cabecalho is None:
            return jsonify({
                "erro": "Não encontrei a coluna de código da filial. Use a planilha baixada pela própria aba Filiais."
            }), 400

        usuario = session.get("username")
        linhas_validas = []
        vistos = set()
        erros = []
        ignoradas = 0
        total_linhas = 0

        # As linhas restantes do iterador começam imediatamente após o cabeçalho.
        for numero_linha, valores in enumerate(linhas_iter, start=cab_row + 1):
            valores = tuple(valores or ())
            if not valores or all(v is None or str(v).strip() == "" for v in valores):
                continue
            total_linhas += 1

            def ler(campo):
                idx = mapa_colunas.get(campo)
                if idx is None or idx >= len(valores):
                    return ""
                valor = valores[idx]
                if valor is None:
                    return ""
                if isinstance(valor, bool):
                    return "1" if valor else "0"
                if isinstance(valor, float) and valor.is_integer():
                    return str(int(valor))
                return str(valor).strip()

            codigo = ler("codigo").strip()
            if not codigo:
                ignoradas += 1
                if len(erros) < 20:
                    erros.append(f"Linha {numero_linha}: código da filial não informado.")
                continue
            if codigo in vistos:
                ignoradas += 1
                if len(erros) < 20:
                    erros.append(f"Linha {numero_linha}: código {codigo} repetido na planilha.")
                continue
            vistos.add(codigo)

            nome_filial = ler("nome")
            cidade = ler("cidade")
            uf = ler("uf").upper().strip()
            bandeira = ler("bandeira").upper().strip()
            status_txt = ler("status").strip()
            previsao_abertura = _data_filial_iso(ler("previsao_abertura"))

            if uf and (len(uf) != 2 or uf not in UF_NOMES):
                ignoradas += 1
                if len(erros) < 20:
                    erros.append(f"Linha {numero_linha}: UF '{uf}' inválida para a filial {codigo}.")
                continue
            if bandeira and bandeira not in ("DSP", "DPA"):
                ignoradas += 1
                if len(erros) < 20:
                    erros.append(f"Linha {numero_linha}: bandeira '{bandeira}' inválida para a filial {codigo}.")
                continue

            status = "" if not status_txt else _status_filial_importacao(status_txt, None)
            if status_txt and status is None:
                ignoradas += 1
                if len(erros) < 20:
                    erros.append(f"Linha {numero_linha}: status '{status_txt}' inválido para a filial {codigo}.")
                continue

            linhas_validas.append({
                "codigo": codigo,
                "nome": nome_filial,
                "cidade": cidade,
                "uf": uf,
                "bandeira": bandeira,
                "status": status,
                "previsao_abertura": previsao_abertura,
            })

        if not linhas_validas:
            return jsonify({"erro": "Nenhuma filial válida foi encontrada na planilha."}), 400

        resultado = db.importar_filiais_em_lote(linhas_validas, usuario)
        criadas = int(resultado.get("criadas") or 0)
        atualizadas = int(resultado.get("atualizadas") or 0)
        sem_alteracao = int(resultado.get("sem_alteracao") or 0)
        processadas = criadas + atualizadas + sem_alteracao
        try:
            db.registrar_importacao("filiais", arquivo.filename, total_linhas, processadas, criadas, atualizadas, ignoradas, usuario, "concluida", f"Sem alteração: {sem_alteracao}")
        except Exception as audit_err:
            print(f"[aviso] Falha ao registrar auditoria de importação: {audit_err}")

        if criadas or atualizadas:
            try:
                db.registrar_movimentacao(
                    0,
                    "importacao_filiais",
                    str(criadas + atualizadas),
                    usuario,
                    f"Importação de filiais: {total_linhas} linha(s) lida(s), {processadas} processada(s), {criadas} criada(s), {atualizadas} atualizada(s), {sem_alteracao} sem alteração e {ignoradas} ignorada(s).",
                    tabela="sistema",
                )
            except Exception as hist_err:
                # O histórico não deve desfazer uma importação que já foi concluída.
                print(f"[aviso] Importação de filiais concluída, mas falhou o histórico: {type(hist_err).__name__}: {hist_err}")

        return jsonify({
            "ok": True,
            "arquivo": arquivo.filename,
            "total_linhas": total_linhas,
            "processadas": processadas,
            "criadas": criadas,
            "atualizadas": atualizadas,
            "sem_alteracao": sem_alteracao,
            "ignoradas": ignoradas,
            "erros": erros,
        })
    except Exception as e:
        print(f"[erro] Falha ao importar filiais: {type(e).__name__}: {e}")
        return jsonify({"erro": f"Erro ao processar a planilha ({type(e).__name__}). Consulte o log do servidor para o detalhe técnico."}), 500


@app.route("/api/filiais", methods=["POST"])
@manager_required
def api_criar_filial():
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip()
    nome = (dados.get("nome") or "").strip()
    cidade = (dados.get("cidade") or "").strip()
    uf = (dados.get("uf") or "").strip().upper()[:2]
    bandeira = (dados.get("bandeira") or "").strip().upper()
    if bandeira not in ("", "DSP", "DPA"):
        return jsonify({"erro": "Bandeira inválida. Use DSP ou DPA."}), 400
    previsao_abertura = _data_filial_iso(dados.get("previsao_abertura"))
    ativo = str(dados.get("ativo", "1") or "1").strip().lower()
    mapa_status = {"ativa":"1","ativo":"1","1":"1","inativa":"0","inativo":"0","0":"0","inaugurar":"inaugurar","pendente":"pendente"}
    ativo = mapa_status.get(ativo, "1")
    if not codigo:
        return jsonify({"erro": "Código da filial é obrigatório."}), 400
    try:
        novo_id = db.criar_filial(codigo, nome, cidade, uf, ativo, session.get("username"), bandeira=bandeira, previsao_abertura=previsao_abertura)
    except Exception:
        return jsonify({"erro": "Já existe uma filial cadastrada com este código."}), 409
    db.registrar_movimentacao(0,"criacao_filial","1",session.get("username"),f"Filial {codigo} criada · status={ativo} · UF={uf} · previsão={previsao_abertura or '-'}",tabela="sistema")
    return jsonify({"ok": True, "id": novo_id}), 201


@app.route("/api/filiais/<int:filial_id>", methods=["PUT"])
@manager_required
def api_atualizar_filial(filial_id):
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip()
    nome = (dados.get("nome") or "").strip()
    cidade = (dados.get("cidade") or "").strip()
    uf = (dados.get("uf") or "").strip().upper()[:2]
    bandeira = (dados.get("bandeira") or "").strip().upper()
    if bandeira not in ("", "DSP", "DPA"):
        return jsonify({"erro": "Bandeira inválida. Use DSP ou DPA."}), 400
    previsao_abertura = _data_filial_iso(dados.get("previsao_abertura"))
    ativo = str(dados.get("ativo", "1") or "1").strip().lower()
    mapa_status = {"ativa":"1","ativo":"1","1":"1","inativa":"0","inativo":"0","0":"0","inaugurar":"inaugurar","pendente":"pendente"}
    ativo = mapa_status.get(ativo, "1")
    if not codigo:
        return jsonify({"erro": "Código da filial é obrigatório."}), 400
    anterior=db.buscar_filial_por_id(filial_id)
    try:
        ok = db.atualizar_filial(filial_id, codigo, nome, cidade, uf, ativo, bandeira=bandeira, previsao_abertura=previsao_abertura)
    except Exception:
        return jsonify({"erro": "Já existe outra filial com este código."}), 409
    if ok:
        resumo_ant=f"status={anterior.get('ativo') if anterior else '-'}; UF={anterior.get('uf') if anterior else '-'}; bandeira={anterior.get('bandeira') if anterior else '-'}; previsão={anterior.get('previsao_abertura') if anterior else '-'}"
        resumo_novo=f"status={ativo}; UF={uf}; bandeira={bandeira or '-'}; previsão={previsao_abertura or '-'}"
        db.registrar_movimentacao(0,"alteracao_filial","1",session.get("username"),f"Filial {codigo} alterada · antes: {resumo_ant} · depois: {resumo_novo}",tabela="sistema")
        return jsonify({"ok": True})
    return jsonify({"erro": "Filial não encontrada."}), 404


@app.route("/api/filiais/<int:filial_id>", methods=["DELETE"])
@manager_required
def api_excluir_filial(filial_id):
    """Exclusão individual no mesmo padrão visual/operacional do Estoque."""
    filial = db.buscar_filial_por_id(filial_id)
    if not filial:
        return jsonify({"erro": "Filial não encontrada."}), 404
    try:
        excluidas, _ = db.excluir_filiais_em_lote([filial_id], desvincular_equipamentos=True)
    except Exception:
        app.logger.exception("Erro ao excluir filial %s", filial_id)
        return jsonify({"erro": "Erro ao excluir filial."}), 500
    if not excluidas:
        return jsonify({"erro": "Erro ao excluir filial."}), 500
    db.registrar_movimentacao(
        0,
        "exclusao_filial",
        "1",
        session.get("username"),
        f"Filial {filial.get('codigo') or filial_id} excluída.",
        tabela="sistema",
    )
    return jsonify({"ok": True, "filial": filial})


@app.route("/api/filiais/excluir-em-lote", methods=["POST"])
@manager_required
def api_excluir_filiais_em_lote():
    """Exclusão em massa seguindo o mesmo fluxo usado no Estoque."""
    dados = request.get_json(force=True) or {}
    ids = dados.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"erro": "Nenhuma filial selecionada."}), 400

    ids_limpos = []
    for valor in ids:
        try:
            filial_id = int(valor)
        except (TypeError, ValueError):
            continue
        if filial_id > 0 and filial_id not in ids_limpos:
            ids_limpos.append(filial_id)
    if not ids_limpos:
        return jsonify({"erro": "Nenhuma filial válida selecionada."}), 400

    # Captura os dados antes da exclusão para registrar no histórico.
    atuais = {int(f["id"]): f for f in db.listar_filiais(incluir_inativas=True) if f.get("id") is not None}
    try:
        excluidas, nao_encontradas = db.excluir_filiais_em_lote(ids_limpos, desvincular_equipamentos=True)
    except Exception:
        app.logger.exception("Erro ao excluir filiais em lote")
        return jsonify({"erro": "Erro ao excluir as filiais selecionadas."}), 500

    for info in excluidas:
        filial = atuais.get(int(info.get("id") or 0), {})
        db.registrar_movimentacao(
            0,
            "exclusao_filial",
            "1",
            session.get("username"),
            f"Filial {filial.get('codigo') or info.get('codigo') or info.get('id')} excluída (exclusão em massa).",
            tabela="sistema",
        )

    return jsonify({
        "ok": True,
        "excluidos": len(excluidas),
        "nao_encontradas": nao_encontradas,
        "build": APP_BUILD,
    })


@app.route("/api/produtos", methods=["GET"])
@login_required
def api_listar_produtos():
    return jsonify(db.listar_produtos())


@app.route("/api/produtos", methods=["POST"])
@manager_required
def api_criar_produto():
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip()
    descricao = (dados.get("descricao") or "").strip()
    try:
        qtde_por_loja = int(dados.get("qtde_por_loja") or 1)
    except (TypeError, ValueError):
        qtde_por_loja = 0
    if not codigo or not descricao:
        return jsonify({"erro": "Código de cadastro e descrição são obrigatórios."}), 400
    if qtde_por_loja < 1:
        return jsonify({"erro": "A quantidade necessária por loja deve ser no mínimo 1."}), 400
    try:
        novo_id = db.criar_produto(codigo, descricao, qtde_por_loja, session.get("username"))
    except Exception:
        return jsonify({"erro": "Já existe um produto cadastrado com este código."}), 409
    return jsonify({"ok": True, "id": novo_id}), 201


@app.route("/api/produtos/<int:produto_id>", methods=["PUT"])
@manager_required
def api_atualizar_produto(produto_id):
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip()
    descricao = (dados.get("descricao") or "").strip()
    try:
        qtde_por_loja = int(dados.get("qtde_por_loja") or 1)
    except (TypeError, ValueError):
        qtde_por_loja = 0
    if not codigo or not descricao:
        return jsonify({"erro": "Código de cadastro e descrição são obrigatórios."}), 400
    if qtde_por_loja < 1:
        return jsonify({"erro": "A quantidade necessária por loja deve ser no mínimo 1."}), 400
    try:
        ok = db.atualizar_produto(produto_id, codigo, descricao, qtde_por_loja)
    except Exception:
        return jsonify({"erro": "Já existe outro produto com este código."}), 409
    return (jsonify({"ok": True}) if ok else (jsonify({"erro": "Produto não encontrado."}), 404))


@app.route("/api/produtos/<int:produto_id>", methods=["DELETE"])
@manager_required
def api_excluir_produto(produto_id):
    ok = db.excluir_produto(produto_id)
    return (jsonify({"ok": True}) if ok else (jsonify({"erro": "Produto não encontrado."}), 404))


# ---------------------------------------------------------------------
# API - Kit padrão de loja
# ---------------------------------------------------------------------

@app.route("/api/configuracao-expansao", methods=["GET"])
@login_required
def api_configuracao_expansao():
    return jsonify({"meta_lojas": db.obter_meta_lojas_expansao()})


@app.route("/api/configuracao-expansao", methods=["PUT"])
@edit_required
def api_salvar_configuracao_expansao():
    dados = request.get_json(force=True) or {}
    try:
        meta_lojas = int(dados.get("meta_lojas"))
    except (TypeError, ValueError):
        return jsonify({"erro": "Informe uma quantidade válida de lojas."}), 400
    if meta_lojas < 1 or meta_lojas > 999:
        return jsonify({"erro": "A quantidade de lojas deve ficar entre 1 e 999."}), 400
    meta_anterior = db.obter_meta_lojas_expansao()
    meta_lojas = db.salvar_meta_lojas_expansao(meta_lojas, session.get("username"))
    if int(meta_anterior) != int(meta_lojas):
        db.registrar_movimentacao(
            0,
            "meta_lojas",
            str(meta_lojas),
            session.get("username"),
            f"Meta do lote de inauguração alterada de {meta_anterior} para {meta_lojas} loja(s).",
            tabela="sistema",
        )
    return jsonify({"ok": True, "meta_lojas": meta_lojas})


@app.route("/api/kit-padrao", methods=["GET"])
@login_required
def api_listar_kit_padrao():
    return jsonify(db.listar_kit_padrao_loja())

@app.route("/api/kit-padrao", methods=["POST"])
@manager_required
def api_criar_item_kit_padrao():
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip() or None
    descricao = (dados.get("descricao") or "").strip()
    try:
        quantidade = int(dados.get("quantidade") or 1)
    except (TypeError, ValueError):
        quantidade = 0
    if not descricao:
        return jsonify({"erro": "Descrição do item é obrigatória."}), 400
    if quantidade < 1:
        return jsonify({"erro": "A quantidade do kit deve ser no mínimo 1."}), 400
    novo_id = db.criar_item_kit(codigo, descricao, quantidade, session.get("username"))
    return jsonify({"ok": True, "id": novo_id}), 201

@app.route("/api/kit-padrao/<int:item_id>", methods=["PUT"])
@manager_required
def api_atualizar_item_kit_padrao(item_id):
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip() or None
    descricao = (dados.get("descricao") or "").strip()
    try:
        quantidade = int(dados.get("quantidade") or 1)
    except (TypeError, ValueError):
        quantidade = 0
    if not descricao:
        return jsonify({"erro": "Descrição do item é obrigatória."}), 400
    if quantidade < 1:
        return jsonify({"erro": "A quantidade do kit deve ser no mínimo 1."}), 400
    ok = db.atualizar_item_kit(item_id, codigo, descricao, quantidade)
    return (jsonify({"ok": True}) if ok else (jsonify({"erro": "Item do kit não encontrado."}), 404))

@app.route("/api/kit-padrao/<int:item_id>", methods=["DELETE"])
@manager_required
def api_excluir_item_kit_padrao(item_id):
    ok = db.excluir_item_kit(item_id)
    return (jsonify({"ok": True}) if ok else (jsonify({"erro": "Item do kit não encontrado."}), 404))


@app.route("/imobilizados")
@login_required
def pagina_imobilizados():
    return render_template(
        "imobilizados.html",
        username=session.get("username"),
        role=session.get("role") or "user",
        is_admin=session.get("role") == "admin",
    )


# ---------------------------------------------------------------------
# API - Imobilizados
# ---------------------------------------------------------------------

@app.route("/api/imobilizados", methods=["GET"])
@login_required
def api_listar_imobilizados():
    return jsonify(db.listar_imobilizados())


@app.route("/api/imobilizados", methods=["POST"])
@edit_required
def api_criar_imobilizado():
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip()
    if not codigo:
        return jsonify({"erro": "Código do item é obrigatório."}), 400

    novo = {
        "codigo": codigo,
        "descricao": (dados.get("descricao") or "").strip(),
        "qtde": dados.get("qtde", ""),
        "localizacao": (dados.get("localizacao") or "").strip(),
        "nf_entrada": (dados.get("nf_entrada") or "").strip(),
        "data_entrada": dados.get("data_entrada") or datetime.now().strftime("%Y-%m-%d"),
        "nf_saida": (dados.get("nf_saida") or "").strip(),
        "data_saida": (dados.get("data_saida") or "").strip(),
        "vd_loja": (dados.get("vd_loja") or "").strip(),
        "filial_destino": (dados.get("filial_destino") or "").strip(),
        "local": (dados.get("local") or "").strip(),
        "armazenagem": (dados.get("armazenagem") or "").strip(),
        "status": (dados.get("status") or "").strip(),
        "nro_imobilizado": (dados.get("nro_imobilizado") or "").strip(),
        "nro_serie": (dados.get("nro_serie") or "").strip(),
        "nro_patrimonio": (dados.get("nro_patrimonio") or "").strip(),
        "tipo_estoque": (dados.get("tipo_estoque") or "").strip(),
        "pedido": (dados.get("pedido") or "").strip(),
        "val_aquis": (dados.get("val_aquis") or "").strip(),
        "chamado": (dados.get("chamado") or "").strip(),
        "criado_por": session.get("username"),
    }
    try:
        qtde_informada = int(float(dados.get("qtde") or 1))
    except (ValueError, TypeError):
        qtde_informada = 1
    qtde_informada = max(qtde_informada, 1)
    linhas = [dict(novo, qtde="1") for _ in range(qtde_informada)]
    total = db.criar_imobilizados_em_lote(linhas, session.get("username"), observacao="Cadastro manual do imobilizado")
    return jsonify({"ok": True, "criados": total, "codigo": codigo}), 201


@app.route("/api/imobilizados/<int:item_id>", methods=["PUT"])
@edit_required
def api_atualizar_imobilizado(item_id):
    dados = request.get_json(force=True)
    item_antes = db.buscar_imobilizado_por_id(item_id)
    if not item_antes:
        return jsonify({"erro": "Item não encontrado."}), 404

    dados["atualizado_por"] = session.get("username")
    dados["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ok = db.atualizar_imobilizado(item_id, dados)
    if not ok:
        return jsonify({"erro": "Item não encontrado."}), 404

    db.registrar_movimentacao(item_id, "edicao", None, session.get("username"),
                               "Dados do imobilizado editados", tabela="imobilizados")
    return jsonify({"ok": True})


@app.route("/api/imobilizados/<int:item_id>", methods=["DELETE"])
@edit_required
def api_excluir_imobilizado(item_id):
    item = db.buscar_imobilizado_por_id(item_id)
    if not item:
        return jsonify({"erro": "Item não encontrado."}), 404
    db.registrar_movimentacao(item_id, "exclusao", item.get("qtde"), session.get("username"),
                               f"Imobilizado {item.get('codigo')} excluído", tabela="imobilizados")
    db.excluir_imobilizado(item_id)
    return jsonify({"ok": True, "item": item})


@app.route("/api/imobilizados/restaurar", methods=["POST"])
@edit_required
def api_restaurar_imobilizado():
    dados = request.get_json(force=True)
    if not dados or not dados.get("id"):
        return jsonify({"erro": "Dados inválidos para restaurar."}), 400
    if db.buscar_imobilizado_por_id(dados["id"]):
        return jsonify({"erro": "Este item já existe (não foi excluído ou já foi restaurado)."}), 400
    db.recriar_imobilizado(dados)
    db.registrar_movimentacao(dados["id"], "restauracao", dados.get("qtde"), session.get("username"),
                               "Exclusão desfeita", tabela="imobilizados")
    return jsonify({"ok": True})


@app.route("/api/imobilizados/excluir-em-lote", methods=["POST"])
@edit_required
def api_excluir_imobilizados_em_lote():
    dados = request.get_json(force=True)
    ids = dados.get("ids") or []
    if not ids:
        return jsonify({"erro": "Nenhum item selecionado."}), 400
    for item_id in ids:
        item = db.buscar_imobilizado_por_id(item_id)
        if item:
            db.registrar_movimentacao(item_id, "exclusao", item.get("qtde"), session.get("username"),
                                       f"Imobilizado {item.get('codigo')} excluído (exclusão em massa)",
                                       tabela="imobilizados")
    total = db.excluir_imobilizados_em_lote(ids)
    return jsonify({"ok": True, "excluidos": total})


@app.route("/api/imobilizados/<int:item_id>/movimentacoes")
@login_required
def api_movimentacoes_imobilizado(item_id):
    return jsonify(db.listar_movimentacoes(item_id, tabela="imobilizados"))


@app.route("/api/imobilizados/<int:item_id>/enviar-estoque", methods=["POST"])
@edit_required
def api_enviar_estoque(item_id):
    total = db.enviar_imobilizado_para_estoque(item_id, session.get("username"))
    if total is None:
        return jsonify({"erro": "Imobilizado não encontrado."}), 404
    return jsonify({"ok": True, "criados_no_estoque": total})


@app.route("/api/imobilizados/enviar-estoque-em-lote", methods=["POST"])
@edit_required
def api_enviar_estoque_em_lote():
    dados = request.get_json(force=True)
    ids = dados.get("ids") or []
    if not ids:
        return jsonify({"erro": "Nenhum item selecionado."}), 400
    total_criados = 0
    total_enviados = 0
    for item_id in ids:
        criados = db.enviar_imobilizado_para_estoque(item_id, session.get("username"))
        if criados is not None:
            total_criados += criados
            total_enviados += 1
    return jsonify({"ok": True, "imobilizados_enviados": total_enviados, "criados_no_estoque": total_criados})


@app.route("/export-imobilizados")
@login_required
def exportar_imobilizados_excel():
    itens = db.listar_imobilizados()
    wb = Workbook()
    ws = wb.active
    ws.title = "Imobilizados"
    colunas = ["ID", "Codigo do item", "Descricao", "Qtde", "Localizacao",
               "NF de entrada", "Data de entrada", "NF de saida",
               "Data de saida", "VD / referencia", "Filial destino", "Local",
               "Armazenagem", "Status", "Nro Imobilizado", "Nro Serie",
               "Nro Patrimonio", "Tipo de Estoque", "Criado por",
               "Ultima alteracao por", "Ultima alteracao em",
               "Pedido", "ValAquis.", "Chamado"]
    ws.append(colunas)
    for it in itens:
        ws.append([
            it["id"], it["codigo"], it["descricao"], it["qtde"], it["localizacao"],
            it["nf_entrada"], it["data_entrada"], it["nf_saida"],
            it["data_saida"], it["vd_loja"], it.get("filial_destino"), it.get("local"),
            it.get("armazenagem"), it.get("status"), it.get("nro_imobilizado"),
            it.get("nro_serie"), it.get("nro_patrimonio"), it.get("tipo_estoque"),
            it.get("criado_por"), it.get("atualizado_por"), it.get("atualizado_em"),
            it.get("pedido"), it.get("val_aquis"), it.get("chamado"),
        ])
    larguras = [8, 18, 32, 8, 18, 18, 16, 18, 16, 18, 22, 12, 14, 12, 16, 16, 16, 18, 14, 16, 16, 14, 12, 14]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome_arquivo = f"imobilizados_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )




@app.route("/export-relatorio-lojas", methods=["POST"])
@login_required
def exportar_relatorio_lojas_excel():
    """Gera um relatório gerencial em Excel com o mesmo resumo usado no PDF."""
    dados = request.get_json(force=True) or {}
    estoque = dados.get("estoque") or []
    faltantes = dados.get("faltantes") or []
    kit = dados.get("kit") or []
    meta_lojas = int(dados.get("meta_lojas") or 10)
    lote_pronto = bool(dados.get("lote_pronto"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    # Paleta e estilos
    cor_escura = "1A2029"
    cor_azul = "2876BE"
    cor_verde = "2B915D"
    cor_laranja = "CD8018"
    cor_clara = "F5F7FA"
    cor_cinza = "66707D"
    borda = Border(
        left=Side(style="thin", color="E1E5EA"),
        right=Side(style="thin", color="E1E5EA"),
        top=Side(style="thin", color="E1E5EA"),
        bottom=Side(style="thin", color="E1E5EA"),
    )

    def titulo_planilha(sheet, titulo, subtitulo=None, col_final=5):
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_final)
        c = sheet.cell(1, 1, titulo)
        c.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=cor_escura)
        c.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30
        if subtitulo:
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_final)
            c2 = sheet.cell(2, 1, subtitulo)
            c2.font = Font(name="Aptos", size=9, color=cor_cinza)
            c2.alignment = Alignment(vertical="center")
            sheet.row_dimensions[2].height = 20

    def cabecalho(sheet, row, titulos, fill=cor_escura):
        for col, value in enumerate(titulos, start=1):
            cell = sheet.cell(row, col, value)
            cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borda
        sheet.row_dimensions[row].height = 24

    def ajustar_larguras(sheet, larguras):
        for idx, largura in enumerate(larguras, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = largura

    gerado = dados.get("gerado_em") or datetime.now().isoformat()
    titulo_planilha(ws, "Relatório de Estoque e Capacidade de Lojas", f"Gerado em: {gerado}", 5)
    ajustar_larguras(ws, [26, 28, 25, 25, 36])

    resumo = [
        ("Lojas completas possíveis", dados.get("lojas_possiveis", 0), cor_verde),
        ("Meta de inauguração", f"{meta_lojas} lojas", cor_verde if lote_pronto else cor_laranja),
        ("Unidades em estoque", dados.get("total_unidades", 0), cor_escura),
        ("Produtos no estoque", dados.get("total_produtos", 0), cor_escura),
        ("Categorias com falta", dados.get("categorias_faltantes", 0), cor_laranja),
    ]
    linha = 4
    for rotulo, valor, cor in resumo:
        ws.cell(linha, 1, rotulo).font = Font(name="Aptos", size=10, bold=True, color=cor_cinza)
        ws.cell(linha, 2, valor).font = Font(name="Aptos Display", size=15, bold=True, color=cor)
        ws.cell(linha, 1).fill = PatternFill("solid", fgColor=cor_clara)
        ws.cell(linha, 2).fill = PatternFill("solid", fgColor=cor_clara)
        ws.cell(linha, 1).border = ws.cell(linha, 2).border = borda
        linha += 1

    linha += 1
    ws.cell(linha, 1, "Item(ns) limitante(s)").font = Font(bold=True, color=cor_laranja)
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=5)
    ws.cell(linha, 2, dados.get("item_limitante") or "-").alignment = Alignment(wrap_text=True)
    linha += 2
    ws.cell(linha, 1, "Resumo da simulação").font = Font(bold=True, color=cor_escura)
    ws.merge_cells(start_row=linha+1, start_column=1, end_row=linha+3, end_column=5)
    ws.cell(linha+1, 1, dados.get("detalhe") or "-").alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(linha+1, 1).fill = PatternFill("solid", fgColor=cor_clara)
    ws.cell(linha+1, 1).border = borda
    ws.freeze_panes = "A4"

    # Estoque detalhado
    ws_e = wb.create_sheet("Estoque detalhado")
    titulo_planilha(ws_e, "Estoque detalhado", "Quantidade consolidada por código e descrição", 3)
    cabecalho(ws_e, 4, ["Código", "Descrição do produto", "Qtd. em estoque"], cor_escura)
    for r_idx, item in enumerate(estoque, start=5):
        valores = [item.get("codigo") or "-", item.get("descricao") or "", item.get("quantidade") or 0]
        for c_idx, valor in enumerate(valores, start=1):
            c = ws_e.cell(r_idx, c_idx, valor)
            c.border = borda
            c.alignment = Alignment(vertical="center", wrap_text=(c_idx == 2), horizontal="center" if c_idx == 3 else "left")
            if r_idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F8FAFC")
    ajustar_larguras(ws_e, [22, 52, 20])
    ws_e.freeze_panes = "A5"
    ws_e.auto_filter.ref = f"A4:C{max(4, ws_e.max_row)}"

    # Faltantes
    ws_f = wb.create_sheet(f"Faltantes meta {meta_lojas} lojas")
    titulo_planilha(ws_f, f"Itens faltantes para completar a premissa de {meta_lojas} lojas", "Premissa padrão usada na preparação das inaugurações", 5)
    cabecalho(ws_f, 4, ["Código", "Item", "Em estoque", "Necessário total", "Faltam"], cor_laranja)
    if faltantes:
        for r_idx, item in enumerate(faltantes, start=5):
            valores = [item.get("codigo") or "-", item.get("descricao") or "", item.get("em_estoque") or 0, item.get("necessario_total") or 0, item.get("faltam") or 0]
            for c_idx, valor in enumerate(valores, start=1):
                c = ws_f.cell(r_idx, c_idx, valor)
                c.border = borda
                c.alignment = Alignment(vertical="center", wrap_text=(c_idx == 2), horizontal="center" if c_idx >= 3 else "left")
                if c_idx == 5:
                    c.font = Font(bold=True, color="B43C2D")
                if r_idx % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="FFF8ED")
    else:
        ws_f.merge_cells("A5:E6")
        ws_f["A5"] = f"Premissa atendida: nenhum item faltante para o lote padrão de {meta_lojas} lojas."
        ws_f["A5"].font = Font(bold=True, color=cor_verde)
        ws_f["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ajustar_larguras(ws_f, [22, 48, 18, 20, 16])
    ws_f.freeze_panes = "A5"
    if faltantes:
        ws_f.auto_filter.ref = f"A4:E{ws_f.max_row}"

    # Kit padrão
    ws_k = wb.create_sheet("Kit padrao por loja")
    titulo_planilha(ws_k, "Kit padrão por loja", f"Base por loja utilizada na premissa de inauguração de {meta_lojas} lojas", 5)
    cabecalho(ws_k, 4, ["Código", "Item", "Qtd./loja", "Em estoque", "Lojas suportadas"], cor_azul)
    for r_idx, item in enumerate(kit, start=5):
        valores = [item.get("codigo") or "-", item.get("descricao") or "", item.get("qtd_por_loja") or 0, item.get("em_estoque") or 0, item.get("lojas_suportadas") or 0]
        for c_idx, valor in enumerate(valores, start=1):
            c = ws_k.cell(r_idx, c_idx, valor)
            c.border = borda
            c.alignment = Alignment(vertical="center", wrap_text=(c_idx == 2), horizontal="center" if c_idx >= 3 else "left")
            if c_idx == 5:
                c.font = Font(bold=True, color=cor_azul)
            if r_idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F5F9FD")
    ajustar_larguras(ws_k, [22, 48, 16, 18, 20])
    ws_k.freeze_panes = "A5"
    if kit:
        ws_k.auto_filter.ref = f"A4:E{ws_k.max_row}"

    # Configurações de impressão
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.oddFooter.center.text = "© 2026 · Developed by Alexandre Martins"
        sheet.oddFooter.right.text = "Página &P de &N"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome_arquivo = f"relatorio-estoque-lojas-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------
# API - Itens
# ---------------------------------------------------------------------

@app.route("/api/itens", methods=["GET"])
@login_required
def api_listar():
    return jsonify(db.listar_itens())


@app.route("/api/itens", methods=["POST"])
@edit_required
def api_criar():
    dados = request.get_json(force=True)
    codigo = (dados.get("codigo") or "").strip()
    if not codigo:
        return jsonify({"erro": "Código do item é obrigatório."}), 400

    base = {
        "codigo": codigo,
        "descricao": (dados.get("descricao") or "").strip(),
        "qtde": "1",
        "localizacao": (dados.get("localizacao") or "").strip(),
        "nf_entrada": (dados.get("nf_entrada") or "").strip(),
        "data_entrada": dados.get("data_entrada") or datetime.now().strftime("%Y-%m-%d"),
        "nf_saida": (dados.get("nf_saida") or "").strip(),
        "data_saida": (dados.get("data_saida") or "").strip(),
        "vd_loja": (dados.get("vd_loja") or "").strip(),
        "filial_destino": (dados.get("filial_destino") or "").strip(),
        "local": (dados.get("local") or "").strip(),
        "armazenagem": (dados.get("armazenagem") or "").strip(),
        "status": (dados.get("status") or "").strip(),
        "nro_imobilizado": (dados.get("nro_imobilizado") or "").strip(),
        "nro_serie": (dados.get("nro_serie") or "").strip(),
        "nro_patrimonio": (dados.get("nro_patrimonio") or "").strip(),
        "tipo_estoque": (dados.get("tipo_estoque") or "").strip(),
        "pedido": (dados.get("pedido") or "").strip(),
        "val_aquis": (dados.get("val_aquis") or "").strip(),
        "chamado": (dados.get("chamado") or "").strip(),
        "criado_por": session.get("username"),
    }

    # Cada unidade vira uma linha própria no Estoque (ex: qtde 40 = 40 linhas,
    # cada uma com qtde 1) — isso permite dar saída/retirar item por item.
    try:
        qtde_informada = int(float(dados.get("qtde") or 1))
    except (ValueError, TypeError):
        qtde_informada = 1
    qtde_informada = max(qtde_informada, 1)

    linhas = [dict(base) for _ in range(qtde_informada)]
    total = db.criar_itens_em_lote(linhas, session.get("username"), observacao="Cadastro manual do item")

    return jsonify({"ok": True, "criados": total, "codigo": codigo}), 201


@app.route("/api/itens/<int:item_id>", methods=["PUT"])
@edit_required
def api_atualizar(item_id):
    dados = request.get_json(force=True)
    item_antes = db.buscar_item_por_id(item_id)
    if not item_antes:
        return jsonify({"erro": "Item não encontrado."}), 404

    dados["atualizado_por"] = session.get("username")
    dados["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ok = db.atualizar_item(item_id, dados)
    if not ok:
        return jsonify({"erro": "Item não encontrado."}), 404

    # Registra a movimentação no histórico, se a quantidade mudou.
    if "qtde" in dados:
        try:
            qtde_antes = float(item_antes.get("qtde") or 0)
            qtde_depois = float(dados["qtde"] or 0)
        except ValueError:
            qtde_antes = qtde_depois = None
        if qtde_antes is not None and qtde_depois < qtde_antes:
            diferenca = qtde_antes - qtde_depois
            if dados.get("nf_saida"):
                obs = f"Saída registrada (NF {dados.get('nf_saida')}, destino: {dados.get('filial_destino') or dados.get('vd_loja') or '-'})"
                tipo_mov = "saida"
            else:
                obs = "Retirada de estoque"
                tipo_mov = "retirada"
            db.registrar_movimentacao(item_id, tipo_mov, str(diferenca), session.get("username"), obs)
        elif qtde_antes is not None and qtde_depois > qtde_antes:
            db.registrar_movimentacao(item_id, "ajuste", str(qtde_depois - qtde_antes),
                                       session.get("username"), "Quantidade aumentada manualmente")
        else:
            db.registrar_movimentacao(item_id, "edicao", None, session.get("username"), "Dados do item editados")
    else:
        db.registrar_movimentacao(item_id, "edicao", None, session.get("username"), "Dados do item editados")

    return jsonify({"ok": True})


@app.route("/api/itens/<int:item_id>", methods=["DELETE"])
@edit_required
def api_excluir(item_id):
    item = db.buscar_item_por_id(item_id)
    if not item:
        return jsonify({"erro": "Item não encontrado."}), 404
    db.registrar_movimentacao(item_id, "exclusao", item.get("qtde"), session.get("username"),
                               f"Item {item.get('codigo')} excluído")
    db.excluir_item(item_id)
    return jsonify({"ok": True, "item": item})


@app.route("/api/itens/restaurar", methods=["POST"])
@edit_required
def api_restaurar():
    dados = request.get_json(force=True)
    if not dados or not dados.get("id"):
        return jsonify({"erro": "Dados inválidos para restaurar."}), 400
    if db.buscar_item_por_id(dados["id"]):
        return jsonify({"erro": "Este item já existe (não foi excluído ou já foi restaurado)."}), 400
    db.recriar_item(dados)
    db.registrar_movimentacao(dados["id"], "restauracao", dados.get("qtde"), session.get("username"),
                               "Exclusão desfeita")
    return jsonify({"ok": True})


@app.route("/api/itens/excluir-em-lote", methods=["POST"])
@edit_required
def api_excluir_em_lote():
    dados = request.get_json(force=True)
    ids = dados.get("ids") or []
    if not ids:
        return jsonify({"erro": "Nenhum item selecionado."}), 400
    for item_id in ids:
        item = db.buscar_item_por_id(item_id)
        if item:
            db.registrar_movimentacao(item_id, "exclusao", item.get("qtde"), session.get("username"),
                                       f"Item {item.get('codigo')} excluído (exclusão em massa)")
    total = db.excluir_itens_em_lote(ids)
    return jsonify({"ok": True, "excluidos": total})


@app.route("/api/itens/<int:item_id>/movimentacoes")
@login_required
def api_movimentacoes(item_id):
    return jsonify(db.listar_movimentacoes(item_id, tabela="itens"))


@app.route("/export")
@login_required
def exportar_excel():
    itens = db.listar_itens()
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    colunas = ["ID", "Codigo do item", "Descricao", "Qtde", "Localizacao",
               "NF de entrada", "Data de entrada", "NF de saida",
               "Data de saida", "VD / referencia", "Filial destino", "Local",
               "Armazenagem", "Status", "Nro Imobilizado", "Nro Serie",
               "Nro Patrimonio", "Tipo de Estoque", "Criado por",
               "Ultima alteracao por", "Ultima alteracao em",
               "Pedido", "ValAquis.", "Chamado"]
    ws.append(colunas)
    for it in itens:
        ws.append([
            it["id"], it["codigo"], it["descricao"], it["qtde"], it["localizacao"],
            it["nf_entrada"], it["data_entrada"], it["nf_saida"],
            it["data_saida"], it["vd_loja"], it.get("filial_destino"), it.get("local"),
            it.get("armazenagem"), it.get("status"), it.get("nro_imobilizado"),
            it.get("nro_serie"), it.get("nro_patrimonio"), it.get("tipo_estoque"),
            it.get("criado_por"), it.get("atualizado_por"), it.get("atualizado_em"),
            it.get("pedido"), it.get("val_aquis"), it.get("chamado"),
        ])
    larguras = [8, 18, 32, 8, 18, 18, 16, 18, 16, 18, 22, 12, 14, 12, 16, 16, 16, 18, 14, 16, 16, 14, 12, 14]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome_arquivo = f"estoque_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------
# Importação de planilha Excel
# ---------------------------------------------------------------------

def _normalizar(texto):
    """Deixa o texto minúsculo, sem acento e sem espaços/pontuação, para
    comparar nomes de coluna de forma tolerante (ex: 'Nº Patrimônio' == 'nro patrimonio')."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = "".join(c for c in texto if c.isalnum())
    return texto


TIPOS_ESTOQUE_CANONICOS = ["Expansão", "Sustentação", "Requalificação", "Ampliação", "Realocação", "Reversa"]


def _canonicalizar_tipo_estoque(valor):
    """Se o valor bater (ignorando acento/maiúscula) com um dos tipos de estoque
    oficiais, devolve a grafia oficial. Caso contrário, devolve o texto original."""
    if not valor:
        return valor
    norm = _normalizar(valor)
    for tipo in TIPOS_ESTOQUE_CANONICOS:
        if _normalizar(tipo) == norm:
            return tipo
    return valor


# Cada campo do sistema aceita várias variações possíveis de nome de coluna
# na planilha (já normalizadas: sem acento, sem espaço, minúsculo).
ALIASES_COLUNAS = {
    "codigo": ["codigo", "codigodoitem"],
    "descricao": ["descricao", "descricaodoequipamento"],
    "qtde": ["qtde", "quantidade", "qtd"],
    "localizacao": ["localizacao"],
    "nf_entrada": ["nfdeentrada", "nfentrada", "notafiscaldeentrada", "nf"],
    "data_entrada": ["datadeentrada", "dataentrada"],
    "nf_saida": ["nfdesaida", "nfsaida", "notafiscaldesaida"],
    "data_saida": ["datadesaida", "datasaida"],
    "vd_loja": ["vddalojadestino", "vddaloja", "vdloja", "vd", "lojadestino"],
    "filial_destino": ["filialdestino", "filial", "codigofilial", "lojafilial", "destinofilial"],
    "local": ["local"],
    "armazenagem": ["armazenagem", "localarmazenagem"],
    "status": ["status"],
    "nro_imobilizado": ["nroimobilizado", "numeroimobilizado", "imobilizado"],
    "nro_serie": ["nroserie", "numerodeserie", "nserie", "serie"],
    "nro_patrimonio": ["nropatrimonio", "numeropatrimonio", "patrimonio"],
    "tipo_estoque": ["tipodeestoque", "tipoestoque"],
    "pedido": ["pedido"],
    "val_aquis": ["valaquis", "valoraquisicao", "valordeaquisicao"],
    "chamado": ["chamado"],
}


def _mapear_colunas(linha_cabecalho):
    """Recebe a primeira linha da planilha (os títulos das colunas) e devolve
    um dicionário {indice_da_coluna: campo_do_sistema}."""
    mapa = {}
    for indice, titulo in enumerate(linha_cabecalho):
        normalizado = _normalizar(titulo)
        for campo, apelidos in ALIASES_COLUNAS.items():
            if normalizado in apelidos:
                mapa[indice] = campo
                break
    return mapa


def _valor_para_texto(valor):
    """Converte o valor de uma célula do Excel (que pode vir como data,
    número, etc.) para texto simples, do jeito que o sistema espera."""
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return str(valor).strip()


@app.route("/api/itens/importar/validar", methods=["POST"])
@edit_required
def api_validar_importacao_itens():
    tabela_destino=request.form.get("tabela","estoque")
    if tabela_destino not in ("estoque","imobilizados"): tabela_destino="estoque"
    arquivo=request.files.get("arquivo")
    if not arquivo or not arquivo.filename:return jsonify({"erro":"Nenhum arquivo enviado."}),400
    try:
        wb=load_workbook(arquivo,read_only=True,data_only=True); ws=wb.active; linhas=ws.iter_rows(values_only=True)
        try:cabecalho=next(linhas)
        except StopIteration:return jsonify({"erro":"A planilha está vazia."}),400
        mapa=_mapear_colunas(cabecalho)
        if "codigo" not in mapa.values():return jsonify({"erro":"Não encontrei uma coluna de Código do item."}),400
        total=validas=ignoradas=registros=0; erros=[]; amostra=[]
        for n,linha in enumerate(linhas,start=2):
            if linha is None or all(v is None for v in linha):continue
            total+=1; dados={}
            for indice,campo in mapa.items():
                if indice<len(linha):dados[campo]=_valor_para_texto(linha[indice])
            if not dados.get("codigo"):
                ignoradas+=1
                if len(erros)<15:erros.append(f"Linha {n}: código não informado.")
                continue
            validas+=1
            qtd=1
            try:qtd=max(1,int(float(dados.get("qtde") or 1)))
            except Exception:qtd=1
            registros += qtd if tabela_destino=="estoque" else 1
            if len(amostra)<8:amostra.append({"codigo":dados.get("codigo"),"descricao":dados.get("descricao",""),"qtde":qtd,"tipo":dados.get("tipo_estoque","")})
        return jsonify({"ok":True,"arquivo":arquivo.filename,"tabela":tabela_destino,"total_linhas":total,"validas":validas,"ignoradas":ignoradas,"registros_previstos":registros,"erros":erros,"amostra":amostra})
    except Exception as e:
        return jsonify({"erro":f"Erro ao validar a planilha: {e}"}),500

@app.route("/api/itens/importar", methods=["POST"])
@edit_required
def api_importar():
    senha = request.form.get("senha", "")
    usuario_atual = db.buscar_usuario_por_id(session["user_id"])
    if not usuario_atual or not check_password_hash(usuario_atual["password_hash"], senha):
        return jsonify({"erro": "Senha incorreta."}), 403

    tabela_destino = request.form.get("tabela", "estoque")
    if tabela_destino not in ("estoque", "imobilizados"):
        tabela_destino = "estoque"

    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400
    if not arquivo.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"erro": "Envie um arquivo Excel (.xlsx)."}), 400

    try:
        try:
            wb = load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return jsonify({"erro": "Não consegui abrir esse arquivo. Confirme se é um .xlsx válido."}), 400

        if ws is None:
            return jsonify({"erro": "A planilha não tem nenhuma aba com dados."}), 400

        linhas = ws.iter_rows(values_only=True)
        try:
            cabecalho = next(linhas)
        except StopIteration:
            return jsonify({"erro": "A planilha está vazia."}), 400

        mapa_colunas = _mapear_colunas(cabecalho)
        if "codigo" not in mapa_colunas.values():
            return jsonify({"erro": "Não encontrei uma coluna de 'Código do item' na planilha. "
                                     "Verifique se a primeira linha tem os títulos das colunas."}), 400

        usuario = session.get("username")
        novos_itens = []
        ignoradas = 0
        linhas_validas_count = 0
        total_linhas_planilha = 0

        for linha in linhas:
            if linha is None or all(v is None for v in linha):
                continue
            total_linhas_planilha += 1
            dados = {}
            for indice, campo in mapa_colunas.items():
                if indice < len(linha):
                    dados[campo] = _valor_para_texto(linha[indice])
            if not dados.get("codigo"):
                ignoradas += 1
                continue
            linhas_validas_count += 1
            if dados.get("tipo_estoque"):
                dados["tipo_estoque"] = _canonicalizar_tipo_estoque(dados["tipo_estoque"])
            dados["criado_por"] = usuario
            if not dados.get("data_entrada"):
                dados["data_entrada"] = datetime.now().strftime("%Y-%m-%d")

            if tabela_destino == "estoque":
                # Cada unidade da planilha vira uma linha própria no Estoque
                # (ex: qtde 40 numa linha da planilha = 40 linhas no sistema).
                try:
                    qtde_linha = int(float(dados.get("qtde") or 1))
                except (ValueError, TypeError):
                    qtde_linha = 1
                qtde_linha = max(qtde_linha, 1)
                base = dict(dados)
                base["qtde"] = "1"
                novos_itens.extend(dict(base) for _ in range(qtde_linha))
            else:
                # Imobilizado: mantém a quantidade exatamente como veio na planilha,
                # numa única linha (não é dividido).
                novos_itens.append(dados)

        if not novos_itens:
            return jsonify({"erro": "Nenhuma linha válida encontrada (confira se a coluna 'Código do item' está preenchida)."}), 400

        if tabela_destino == "estoque":
            total = db.criar_itens_em_lote(
                novos_itens, usuario,
                observacao=f"Importado via planilha ({arquivo.filename})"
            )
        else:
            total = db.criar_imobilizados_em_lote(
                novos_itens, usuario,
                observacao=f"Importado via planilha ({arquivo.filename})"
            )

        try:
            db.registrar_importacao(tabela_destino, arquivo.filename, total_linhas=total_linhas_planilha, validas=linhas_validas_count, criadas=total, atualizadas=0, ignoradas=ignoradas, usuario=usuario, status="concluida")
        except Exception as audit_err:
            print(f"[aviso] Falha ao registrar auditoria de importação: {audit_err}")
        return jsonify({"ok": True, "importados": total, "ignoradas": ignoradas, "tabela": tabela_destino})

    except Exception as e:
        print(f"[erro] Falha ao importar planilha: {e}")
        return jsonify({"erro": f"Erro ao processar a planilha: {e}"}), 500


# ---------------------------------------------------------------------
# API - Usuários (só admin)
# ---------------------------------------------------------------------

@app.route("/api/usuarios", methods=["POST"])
@admin_required
def api_criar_usuario():
    dados = request.get_json(force=True)
    username = (dados.get("username") or "").strip()
    password = dados.get("password") or ""
    role = dados.get("role") if dados.get("role") in ("admin", "gestor", "operador", "consulta", "user") else "operador"

    if not username or not password:
        return jsonify({"erro": "Usuário e senha são obrigatórios."}), 400
    if len(password) < 6:
        return jsonify({"erro": "A senha precisa ter pelo menos 6 caracteres."}), 400
    if db.buscar_usuario_por_username(username):
        return jsonify({"erro": "Já existe um usuário com esse nome."}), 400

    db.criar_usuario(username, password, role)
    # A senha temporária é devolvida somente nesta resposta ao Administrador.
    # No banco permanece apenas o hash; não há recuperação posterior em texto aberto.
    return jsonify({"ok": True, "username": username, "senha_temporaria": password}), 201


@app.route("/api/usuarios/<int:user_id>/forcar-troca-senha", methods=["POST"])
@admin_required
def api_forcar_troca_senha(user_id):
    alvo = db.buscar_usuario_por_id(user_id)
    if not alvo:
        return jsonify({"erro": "Usuário não encontrado."}), 404
    db.forcar_troca_senha(user_id)
    return jsonify({"ok": True})


@app.route("/api/usuarios/<int:user_id>/reset-mfa", methods=["POST"])
@admin_required
def api_reset_mfa_usuario(user_id):
    if not _csrf_ok():
        return jsonify({"erro": "Token de segurança inválido."}), 400
    alvo = db.buscar_usuario_por_id(user_id)
    if not alvo:
        return jsonify({"erro": "Usuário não encontrado."}), 404
    if alvo.get("mfa_enabled") != "1":
        return jsonify({"erro": "Este usuário não possui MFA ativo."}), 400
    db.desativar_mfa_usuario(user_id)
    db.registrar_evento_login(alvo.get("username"), _client_ip(), "mfa_reset_admin", f"MFA resetado pelo administrador {session.get('username')}")
    return jsonify({"ok": True, "mensagem": "MFA resetado. O usuário deverá configurar novamente no próximo acesso."})


@app.route("/api/usuarios/<int:user_id>", methods=["DELETE"])
@admin_required
def api_excluir_usuario(user_id):
    if user_id == session.get("user_id"):
        return jsonify({"erro": "Você não pode excluir o próprio usuário enquanto está logado com ele."}), 400

    alvo = db.buscar_usuario_por_id(user_id)
    if alvo and alvo["role"] == "admin" and db.contar_admins() <= 1:
        return jsonify({"erro": "Precisa existir pelo menos um administrador."}), 400

    ok = db.excluir_usuario(user_id)
    if not ok:
        return jsonify({"erro": "Usuário não encontrado."}), 404
    return jsonify({"ok": True})


# ---------------------------------------------------------------------
# Inicialização / execução
# ---------------------------------------------------------------------

db.init_db()


def descobrir_ip_local():
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "SEU-IP-LOCAL"


if __name__ == "__main__":
    ip = descobrir_ip_local()
    print("=" * 60)
    print("Acesse neste computador em:  http://localhost:5000")
    print(f"Outras pessoas na mesma rede acessam em:  http://{ip}:5000")
    print("=" * 60)

    try:
        from waitress import serve
        serve(app, host="0.0.0.0", port=5000)
    except ImportError:
        print("\n[Aviso] 'waitress' não instalado — rodando com o servidor")
        print("de desenvolvimento do Flask.\n")
        app.run(host="0.0.0.0", port=5000, debug=False)
